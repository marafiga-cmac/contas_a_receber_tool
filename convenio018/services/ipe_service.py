"""Processamento de faturamento para convênio Ipê."""

from __future__ import annotations

from io import BytesIO
import pandas as pd

from ..utils.normalizers import _dedupe_headers, _drop_nan_only_columns, _drop_rows_without_inicio
from ..utils.parsers import _excel_col_idx


import re
import pdfplumber

def _converter_valor(valor_str: str) -> float:
    """Converte string de valor financeiro (BR) para float."""
    v = valor_str.replace('.', '').replace(',', '.')
    try:
        return float(v)
    except ValueError:
        return 0.0

def extrair_dados_demonstrativo_ipe(pdf_file) -> dict:
    """
    Lê o PDF de Demonstrativo de Pagamentos do IPÊ.
    Extrai:
    - Data de crédito
    - Totais financeiros (Total no Processo, IRF Retido, Líquido a receber)
    - Tabela de documentos (Nro Doc e Valor Pago)
    Retorna um dicionário estruturado.
    """
    texto_completo = ""
    
    try:
        # Garante que o buffer esteja no início caso tenha sido lido antes
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                texto_pagina = page.extract_text()
                if texto_pagina:
                    texto_completo += texto_pagina + "\n"
    except Exception as e:
        raise ValueError(f"Não foi possível processar o arquivo PDF: {str(e)}")

    # 1. Busca de Metadados (Texto Consolidado para evitar quebras de linha)
    texto_para_busca = texto_completo.replace('\n', ' ')
    
    # Data de Crédito
    m_credito = re.search(r"(?i)Creditado\s+em\s*[:\s]*(\d{2}/\d{2}/\d{4})", texto_para_busca)
    data_credito = m_credito.group(1) if m_credito else None
    
    if not data_credito:
        raise ValueError("Não foi possível encontrar a 'Data de Crédito'. O PDF pode não ser um demonstrativo válido do Ipê ou o padrão mudou.")

    # Totais Financeiros
    def buscar_total(label_regex):
        match = re.search(label_regex + r"\s*[:\s]*([\d\.,]+)", texto_para_busca, re.IGNORECASE)
        return _converter_valor(match.group(1)) if match else 0.0

    total_processo = buscar_total(r"Total\s+no\s+Processo")
    irf_retido = buscar_total(r"IRF\s+Retido")
    liquido_receber = buscar_total(r"L[íi]quido\s+a\s+receber")

    # 2. Tabela de Documentos
    documentos = []
    # Novo padrão sugerido: captura Grupo 1 (Doc) e Grupo 2 (Valor no final $)
    # O padrão (\d{2}\.?\d{3}) captura formatos como 50.002 ou 50002
    doc_pattern = re.compile(r"(\d{2}\.?\d{3})\s+.*?([\d\.]*,\d{2})$")
    
    for linha in texto_completo.split('\n'):
        linha_clean = linha.strip()
        if not linha_clean:
            continue
            
        m_doc = doc_pattern.search(linha_clean)
        if m_doc:
            # Grupo 1: Nro Doc (limpa o ponto)
            nro_doc = m_doc.group(1).replace('.', '')
            # Grupo 2: Valor Pago (converte para float para manter o DF tipado)
            valor_pago = _converter_valor(m_doc.group(2))
            
            documentos.append({"Nro Doc": nro_doc, "Valor Pago": valor_pago})

    df_docs = pd.DataFrame(documentos)
    if not df_docs.empty:
        df_docs["Nro Doc"] = df_docs["Nro Doc"].astype(str)
        df_docs["Valor Pago"] = df_docs["Valor Pago"].astype(float)
    else:
        df_docs = pd.DataFrame(columns=["Nro Doc", "Valor Pago"])

    return {
        "data_credito": data_credito,
        "totais": {
            "total_processo": total_processo,
            "irf_retido": irf_retido,
            "liquido_receber": liquido_receber
        },
        "df_documentos": df_docs
    }
def extrair_detalhado_consultas_ipe(pdf_file):
    text = ""
    # Evita problemas de leitura caso o ponteiro não esteja no início
    pdf_file.seek(0)
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            extr = page.extract_text()
            if extr:
                text += extr + " "
    
    # Troca quebras de linha por espaços e remove espaços duplos
    text_limpo = re.sub(r'\s+', ' ', text.replace('\n', ' '))

    # Fatia o texto a cada CPF encontrado (o CPF marca o fim exato de um registro)
    parts = re.split(r'(\d{3}\.\d{3}\.\d{3}-\d{2})', text_limpo)
    
    documentos = []
    
    # Itera de 2 em 2 (bloco de texto + CPF correspondente)
    for i in range(0, len(parts) - 1, 2):
        bloco = parts[i]
        
        # Pega a Matrícula (13 dígitos) que é o "meio" do registro
        match_mat = re.search(r'\b(\d{13})\b', bloco)
        if not match_mat:
            continue
        
        matricula = match_mat.group(1)
        before_mat, after_mat = bloco.split(matricula, 1)
        
        # EXTRAIR NOME (pega as palavras em maiúsculo logo antes da matrícula)
        nome = "NÃO IDENTIFICADO"
        match_nome = re.search(r'(?:^|\s)\d{2,3}\s+([A-ZÀ-Ÿ][A-ZÀ-Ÿ\s]+)$', before_mat)
        if match_nome:
            nome = match_nome.group(1).strip()
        else:
            match_nome_fallback = re.search(r'([A-ZÀ-Ÿ\s]{5,})$', before_mat)
            if match_nome_fallback:
                nome = match_nome_fallback.group(1).strip()
        
        # EXTRAIR STATUS, VLR IPE E N.NOTA
        status_cancelado = False
        if "CANCELADA" in after_mat.upper():
            status_cancelado = True
            vlr_ipe = "CANCELADA"
            n_nota = "-"
        else:
            # Pega Valor IPE e N.Nota que ficam antes do Ref e PINPAD no fim do bloco
            match_valores = re.search(r'(\d+,\d{2})\s+(\d{4,8})\s+\d+\s+[A-Za-z]\s*$', after_mat)
            if match_valores:
                vlr_ipe = match_valores.group(1)
                n_nota = match_valores.group(2)
            else:
                vlr_ipe = "0,00"
                n_nota = "-"
                
        # EXTRAIR HORA E DIA
        hora, dia = "", ""
        match_hora = re.search(r'(\d{2}:\d{2}:\d{2}:\d{1,2})', after_mat)
        if match_hora:
            hora = match_hora.group(1)
            
        after_mat_no_hour = after_mat.replace(hora, '') if hora else after_mat
        match_dia = re.search(r'(?:^|\s)(\d{2})(?=\s)', after_mat_no_hour)
        if match_dia:
            dia = match_dia.group(1)

        documentos.append({
            "N.Nota": str(n_nota).replace('.', '').strip(),
            "Nome": nome,
            "Dia": dia,
            "Hora": hora,
            "Vlr IPE": vlr_ipe,
            "Status_Cancelado": status_cancelado
        })
        
    return pd.DataFrame(documentos)

def processar_ipe_xls_adicionais(uploaded_files) -> pd.DataFrame:
    """
    Recebe lista de UploadedFile do Streamlit referentes a relatórios internos XLS/XLSX.
    - Suporte a .xls (xlrd) e .xlsx (openpyxl).
    - Extração de metadados das células G5 (Remessa) e X5 (Competência).
    - Correção de cabeçalho na Linha 11: Mover B->C, AB->AF e AQ->AR.
    - Leitura dos dados a partir da Linha 12.
    - Limpeza de colunas vazias e linhas onde a coluna 'Início' esteja nula.
    - Consolidação em um único DataFrame com a coluna 'Arquivo'.
    """
    if not uploaded_files:
        return pd.DataFrame()

    frames = []

    for uf in uploaded_files:
        name = getattr(uf, "name", "arquivo")
        raw = uf.read()
        ext = name.lower().split(".")[-1]

        if ext == "xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
                ws = wb.active

                remessa = ws["G5"].value
                competencia = ws["X5"].value

                row_header = 11
                ws.cell(row_header, _excel_col_idx("C")).value = ws.cell(row_header, _excel_col_idx("B")).value
                ws.cell(row_header, _excel_col_idx("B")).value = None
                ws.cell(row_header, _excel_col_idx("AF")).value = ws.cell(row_header, _excel_col_idx("AB")).value
                ws.cell(row_header, _excel_col_idx("AB")).value = None
                ws.cell(row_header, _excel_col_idx("AR")).value = ws.cell(row_header, _excel_col_idx("AQ")).value
                ws.cell(row_header, _excel_col_idx("AQ")).value = None

                start_col = _excel_col_idx("B")
                max_col = ws.max_column

                headers = []
                for c in range(start_col, max_col + 1):
                    raw_val = ws.cell(row_header, c).value
                    if raw_val is None:
                        col_name = f"COL_{c}"
                    else:
                        s = str(raw_val).strip()
                        if s == "" or s.lower() == "nan":
                            col_name = f"COL_{c}"
                        else:
                            col_name = s
                    headers.append(col_name)

                headers = _dedupe_headers(headers)
                col_indices = list(range(start_col, max_col + 1))

                data_rows = []
                for r in range(row_header + 1, ws.max_row + 1):
                    row_vals = [ws.cell(r, c).value for c in col_indices]
                    if all(v is None or str(v).strip() == "" for v in row_vals):
                        continue
                    data_rows.append(row_vals)

                df = pd.DataFrame(data_rows, columns=headers)
                df.columns = _dedupe_headers(df.columns.tolist())

                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{"Arquivo": name, "Erro": f"Falha ao ler XLSX: {e}"}])
                remessa = None
                competencia = None

        elif ext == "xls":
            try:
                raw_df = pd.read_excel(BytesIO(raw), header=None, engine="xlrd")

                remessa = raw_df.iat[4, _excel_col_idx("G") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("G") - 1) else None
                competencia = raw_df.iat[4, _excel_col_idx("X") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("X") - 1) else None

                hr = 10
                raw_df.iat[hr, _excel_col_idx("C") - 1] = raw_df.iat[hr, _excel_col_idx("B") - 1]
                raw_df.iat[hr, _excel_col_idx("B") - 1] = None
                raw_df.iat[hr, _excel_col_idx("AF") - 1] = raw_df.iat[hr, _excel_col_idx("AB") - 1]
                raw_df.iat[hr, _excel_col_idx("AB") - 1] = None
                raw_df.iat[hr, _excel_col_idx("AR") - 1] = raw_df.iat[hr, _excel_col_idx("AQ") - 1]
                raw_df.iat[hr, _excel_col_idx("AQ") - 1] = None

                start_c = _excel_col_idx("B") - 1
                headers = raw_df.iloc[hr, start_c:].tolist()
                headers = [
                    f"COL_{start_c+i+1}" if (h is None or str(h).strip() == "") else str(h).strip()
                    for i, h in enumerate(headers)
                ]

                df = raw_df.iloc[hr+1:, start_c:].copy()
                headers = _dedupe_headers(headers)
                df.columns = headers
                df = df.dropna(how="all")

                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{"Arquivo": name, "Erro": f"Falha ao ler XLS: {e}"}])
                remessa = None
                competencia = None

        else:
            df = pd.DataFrame([{"Arquivo": name, "Erro": "Extensão não suportada"}])
            remessa = None
            competencia = None

        for meta_col in ["Arquivo", "Remessa (G5)", "Competência (X5)"]:
            if meta_col in df.columns:
                df = df.rename(columns={meta_col: f"{meta_col} (orig)"})

        df.insert(0, "Arquivo", name)
        df.insert(1, "Remessa (G5)", remessa)
        df.insert(2, "Competência (X5)", competencia)

        df.columns = _dedupe_headers(df.columns.tolist())

        df = _drop_nan_only_columns(df)
        df = _drop_rows_without_inicio(df, "Início")

        frames.append(df)

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    out = _drop_nan_only_columns(out)
    out = _drop_rows_without_inicio(out, "Início")

    return out

def limpar_valor(v) -> float:
    """Converte strings financeiras (ex: '80,00' ou 'R$ 80.00') para float cravado."""
    if pd.isna(v) or v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('R$', '').strip()
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0

def nomes_parecidos(n1: str, n2: str) -> bool:
    """
    O 'Match' é verdadeiro SE o primeiro nome for idêntico
    E o tamanho da intersecção entre as duas listas de nomes for >= 2.
    """
    if pd.isna(n1) or pd.isna(n2) or not n1 or not n2:
        return False
    
    parts1 = str(n1).strip().upper().split()
    parts2 = str(n2).strip().upper().split()
    
    if not parts1 or not parts2:
        return False
        
    if parts1[0] != parts2[0]:
        return False
        
    intersection = set(parts1).intersection(set(parts2))
    return len(intersection) >= 2

def executar_identificacao_final(df_fase2: pd.DataFrame, df_fase3: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Cruza Fase 2 (Autorizadas) com Fase 3 (Relatórios Internos XLS).
    Fase 2 tem N.Nota, Nome, Dia, Hora, Vlr IPE, Status_Cancelado.
    Fase 3 tem Arquivo, Remessa (G5), Competência (X5), Início, Término, Nome, Valor, etc.
    """
    if df_fase2 is None or df_fase2.empty or df_fase3 is None or df_fase3.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    encontrados = []
    repetidos = []
    nao_encontrados = []

    fase2_validos = df_fase2[df_fase2.get("Status_Cancelado", False) == False]

    col_nome_f3 = next((c for c in df_fase3.columns if c.strip().lower() in ["nome", "beneficiário", "paciente"]), None)
    col_valor_f3 = next((c for c in df_fase3.columns if c.strip().lower() in ["valor", "vlr", "valor cobrado", "valor pago"]), None)
    col_termino_f3 = next((c for c in df_fase3.columns if c.strip().lower() in ["término", "termino", "data", "data atendimento"]), None)

    if not col_nome_f3 or not col_valor_f3 or not col_termino_f3:
        raise ValueError(f"Faltam colunas essenciais na Fase 3. Encontradas: {df_fase3.columns.tolist()}")

    df_fase3_search = df_fase3.copy()
    df_fase3_search["_valor_limpo"] = df_fase3_search[col_valor_f3].apply(limpar_valor)
    df_fase3_search["_termino_str"] = df_fase3_search[col_termino_f3].astype(str).str.strip()

    for idx, row2 in fase2_validos.iterrows():
        vlr2 = limpar_valor(row2.get("Vlr IPE"))
        nome2 = row2.get("Nome", "")
        dia2 = str(row2.get("Dia", "")).zfill(2)
        n_nota2 = row2.get("N.Nota", "")

        mask_valor = df_fase3_search["_valor_limpo"] == vlr2
        mask_dia = df_fase3_search["_termino_str"].str.startswith(dia2)
        
        candidatos = df_fase3_search[mask_valor & mask_dia]

        matches = []
        for c_idx, cand in candidatos.iterrows():
            if nomes_parecidos(nome2, cand.get(col_nome_f3, "")):
                matches.append(cand)

        if len(matches) == 1:
            match = matches[0]
            encontrados.append({
                "Remessa": match.get("Remessa (G5) (orig)", match.get("Remessa (G5)", "")),
                "N.Nota": n_nota2,
                "Competência": match.get("Competência (X5) (orig)", match.get("Competência (X5)", "")),
                "Data Atendimento": match.get(col_termino_f3, ""),
                "Nome": nome2,
                "Valor": row2.get("Vlr IPE")
            })
        elif len(matches) > 1:
            for match in matches:
                repetidos.append({
                    "Remessa": match.get("Remessa (G5) (orig)", match.get("Remessa (G5)", "")),
                    "N.Nota": n_nota2,
                    "Competência": match.get("Competência (X5) (orig)", match.get("Competência (X5)", "")),
                    "Data Atendimento": match.get(col_termino_f3, ""),
                    "Nome": nome2,
                    "Valor": row2.get("Vlr IPE"),
                    "Status": "Múltiplos registros com mesmo valor e data"
                })
        else:
            nao_encontrados.append({
                "N.Nota": n_nota2,
                "Nome": nome2,
                "Dia": row2.get("Dia"),
                "Vlr IPE": row2.get("Vlr IPE"),
                "Status": "Não localizado no relatório interno"
            })

    df_enc = pd.DataFrame(encontrados, columns=["Remessa", "N.Nota", "Competência", "Data Atendimento", "Nome", "Valor"])
    df_rep = pd.DataFrame(repetidos, columns=["Remessa", "N.Nota", "Competência", "Data Atendimento", "Nome", "Valor", "Status"])
    df_nao = pd.DataFrame(nao_encontrados, columns=["N.Nota", "Nome", "Dia", "Vlr IPE", "Status"])

    return df_enc, df_rep, df_nao
