from __future__ import annotations
import os, re, json
from typing import Any, Dict, List, Optional
from datetime import date, datetime
from io import BytesIO

import pandas as pd


def _drop_nan_only_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas onde todos os valores são vazios/NaN/'nan'."""
    if df is None or df.empty:
        return df

    def _is_empty_cell(x) -> bool:
        if x is None:
            return True
        try:
            if pd.isna(x):
                return True
        except Exception:
            pass
        s = str(x).strip()
        return s == "" or s.lower() == "nan"

    keep = []
    for col in df.columns:
        vals = df[col].tolist()
        keep.append(not all(_is_empty_cell(v) for v in vals))
    return df.loc[:, keep]


def _drop_rows_without_inicio(df: pd.DataFrame, col_name: str = "Início") -> pd.DataFrame:
    """Remove linhas onde a coluna 'Início' está vazia/NaN/'nan'."""
    if df is None or df.empty:
        return df
    # encontra coluna por igualdade case-insensitive
    real_col = None
    tgt = (col_name or "").strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == tgt:
            real_col = c
            break
    if real_col is None:
        return df

    def _has_value(x) -> bool:
        if x is None:
            return False
        try:
            if pd.isna(x):
                return False
        except Exception:
            pass
        s = str(x).strip()
        return not (s == "" or s.lower() == "nan")

    mask = df[real_col].apply(_has_value)
    return df.loc[mask].reset_index(drop=True)


def _dedupe_headers(cols) -> list[str]:
    """Garante nomes únicos de colunas, tratando None/''/'nan' como vazios."""
    used: dict[str, int] = {}
    out: list[str] = []
    for i, c in enumerate(list(cols or [])):
        base = "" if c is None else str(c).strip()
        if base == "" or base.lower() == "nan":
            base = f"COL_{i}"
        n = used.get(base, 0)
        name = f"{base}_{n}" if n > 0 else base
        used[base] = n + 1
        out.append(name)
    return out


# ---- Streamlit é opcional aqui (rodamos sem falhar fora do Streamlit) ----
try:
    import streamlit as st  # para ler session_state se existir
except Exception:  # pragma: no cover
    st = None

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

# ------------------------- Normalização de colunas --------------------------

REQUIRED_FIELDS = {
    "n remessa": "Nº Remessa", "no remessa": "Nº Remessa", "nº remessa": "Nº Remessa", "numero remessa": "Nº Remessa",
    "ref": "Ref.", "ref.": "Ref.", "referencia": "Ref.",
    "n nf": "Nº NF", "no nf": "Nº NF", "nº nf": "Nº NF", "numero nf": "Nº NF", "nfse": "Nº NF",
    "nf recurso": "NF recurso", "nf de recurso": "NF recurso",
    "valor envio xml - remessa": "Valor envio XML - Remessa", "valor envio xml remessa": "Valor envio XML - Remessa",
    "valor pgto": "Valor pgto", "valor pagamento": "Valor pgto",
    "valor glosado": "Valor glosado",
    "imposto": "Imposto", "imp.": "Imposto", "impostos": "Imposto",
    "glosa mantida": "Glosa mantida", "glosa mant.": "Glosa mantida",
    "valor pago": "Valor pago",
    "valor nf": "Valor NF",
    "valor da nf": "Valor NF",
    "vlr nf": "Valor NF",
    "valor nota": "Valor NF",
    "valor nota fiscal": "Valor NF",
}
COMBINED_IMPOSTO_GLOSA_KEYS = ["imposto glosa mantida", "imposto/glosa mantida", "imposto e glosa mantida"]

REMESSAS_COLUMNS = [
    "Nº Remessa","Ref.","Nº NF",
    "Valor envio XML - Remessa","Valor pgto","Valor glosado","Imposto","Glosa mantida"
]
RECURSOS_COLUMNS = [
    "Nº Remessa","Ref.","Nº NF",
    "Valor glosado","Valor pago","Imposto","Glosa mantida"
]

# Colunas por nome (usadas na leitura de datas/valor recursado)
REMESSA_DATE_KEYS = [
    "data pgto remessa", "data remessa", "data pagamento remessa",
    "dt pgto remessa", "dt remessa"
]
RECURSO_DATE_KEYS = [
    "data pgto recurso", "data pagamento recurso",
    "data recurso", "data pgto rec", "dt pgto recurso"
]
VALOR_RECURSADO_KEYS = [
    "valor recursado",  # priorizado
    "valor recurso", "valor recurso (rec)", "valor glosado"
]

CSV_COLS = 13
CSV_DELIM = ";"

# ------------------------------ Helpers base -------------------------------

def _norm(s: Any) -> str:
    import unicodedata
    if s is None: return ""
    s = str(s)
    s2 = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s2 = re.sub(r"\s+", " ", s2)
    return s2.strip().lower()

def _as_date(cell: Any) -> Optional[date]:
    if cell is None or str(cell).strip() == "": return None
    s = str(cell).strip()
    for fmt in ("%d/%m/%Y","%d-%m-%Y","%Y-%m-%d"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    try:
        n = float(s.replace(",", "."))
        base_ord = date(1899,12,30).toordinal()
        return date.fromordinal(base_ord + int(n))
    except Exception:
        return None

def _as_number(cell: Any) -> Optional[float]:
    if cell is None or str(cell).strip() == "": return None
    s = str(cell).strip()
    for sym in ["R$","$"]: s = s.replace(sym,"")
    s = s.replace(" ","")
    if "," in s and "." in s: s = s.replace(".","").replace(",",".")
    elif "," in s: s = s.replace(",",".")
    try: return float(s)
    except ValueError:
        m = re.search(r"-?\d[\d\.,]*", s)
        return _as_number(m.group(0)) if m else None

def _split_imposto_glosa(cell: Any) -> (Optional[float], Optional[float]):
    if cell is None or str(cell).strip() == "": return None, None
    s = str(cell)
    imp = glo = None
    m_imp = re.search(r"imposto[^0-9\-]*(-?\d[\d\.,]*)", s, flags=re.I)
    m_glo = re.search(r"glosa\s*mantida[^0-9\-]*(-?\d[\d\.,]*)", s, flags=re.I)
    if m_imp: imp = _as_number(m_imp.group(1))
    if m_glo: glo = _as_number(m_glo.group(1))
    if imp is None or glo is None:
        nums = re.findall(r"-?\d[\d\.,]*", s)
        if len(nums) >= 2:
            if imp is None: imp = _as_number(nums[0])
            if glo is None: glo = _as_number(nums[1])
    return imp, glo

def _find_col_idx(headers: List[str], candidates: List[str], fallback_idx: Optional[int] = None) -> Optional[int]:
    """Retorna índice da coluna cujo cabeçalho (normalizado) bate com candidates; senão fallback_idx."""
    norm_headers = [_norm(h) for h in headers]
    norm_to_idx = {nh: i for i, nh in enumerate(norm_headers)}
    for cand in candidates:
        nh = _norm(cand)
        if nh in norm_to_idx:
            return norm_to_idx[nh]
    return fallback_idx

# ----------------------------- Google Sheets -------------------------------

def _get_credentials(
    scopes: Optional[List[str]] = None,
    client_secrets_path: str = "client_secret.json",
    token_path: str = "token.json",
) -> Credentials:
    from google.auth.exceptions import RefreshError
    scopes = scopes or SCOPES

    def _save(creds: Credentials):
        os.makedirs(os.path.dirname(token_path) or ".", exist_ok=True)
        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    creds: Optional[Credentials] = None
    if os.path.exists(token_path):
        try:
            creds = Credentials.from_authorized_user_file(token_path, scopes)
        except Exception:
            creds = None

    if creds and not creds.valid and creds.refresh_token:
        try:
            creds.refresh(Request()); _save(creds); return creds
        except (RefreshError, HttpError):
            try: os.remove(token_path)
            except OSError: pass
            creds = None

    if not creds or not creds.valid:
        if not os.path.exists(client_secrets_path):
            raise FileNotFoundError(
                f"Arquivo de client secrets não encontrado: {client_secrets_path}. "
                "Baixe do Google Cloud (OAuth client type: Desktop App)."
            )
        flow = InstalledAppFlow.from_client_secrets_file(client_secrets_path, scopes)
        try:
            creds = flow.run_local_server(
                port=0, open_browser=True,
                authorization_prompt_message="",
                success_message="Autorização concluída. Pode fechar esta aba.",
                access_type="offline", prompt="consent",
            )
        except Exception:
            creds = flow.run_console(
                authorization_prompt_message="Visite a URL, autorize e cole o código aqui:",
                access_type="offline", prompt="consent",
            )
        _save(creds)
    return creds

def _read_sheet_values(service, spreadsheet_id: str, sheet_name: str, header_row: int = 10, **_ignore):
    sheet = service.spreadsheets()
    range_a1 = f"'{sheet_name}'!A{header_row}:ZZ"
    resp = sheet.values().get(spreadsheetId=spreadsheet_id, range=range_a1).execute()
    return resp.get("values", [])

def _build_header_map(headers: List[str]) -> Dict[str, str]:
    return {_norm(h): h for h in headers}

def _is_value_field(out_name: str) -> bool:
    n = _norm(out_name)
    return n.startswith("valor") or "imposto" in n or "glosa mantida" in n

def _extract_required_fields(row_dict: Dict[str, Any], headers_map: Dict[str, str]) -> Dict[str, Any]:
    out_names = {
        "Nº Remessa","Ref.","Nº NF","NF recurso",
        "Valor envio XML - Remessa","Valor pgto","Valor glosado","Imposto","Glosa mantida","Valor pago"
    }
    out: Dict[str, Any] = {k: None for k in out_names}
    for norm_key, out_name in REQUIRED_FIELDS.items():
        if norm_key in headers_map:
            original_header = headers_map[norm_key]
            val = row_dict.get(original_header)
            out[out_name] = _as_number(val) if _is_value_field(out_name) else (val if val != "" else None)

    need_imp = out["Imposto"] is None
    need_glo = out["Glosa mantida"] is None
    if need_imp or need_glo:
        for combined_key in COMBINED_IMPOSTO_GLOSA_KEYS:
            if combined_key in headers_map:
                cel = row_dict.get(headers_map[combined_key])
                imp, glo = _split_imposto_glosa(cel)
                if need_imp and imp is not None: out["Imposto"] = imp; need_imp = False
                if need_glo and glo is not None: out["Glosa mantida"] = glo; need_glo = False
                break
    return out

# --------------------------- Pipeline principal -----------------------------

def processar_convenio_para_json(
    spreadsheet_id: str,
    sheet_name: str,
    data_pagamento: date,
    output_dir: str = ".",
    client_secrets_path: str = "client_secret.json",
    token_path: str = "token.json",
) -> str:
    """
    Lê a planilha e escreve um JSON só com as linhas cuja
    Data pgto remessa == data_pagamento OU Data pgto recurso == data_pagamento.
    Colunas são resolvidas por NOME DE CABEÇALHO (case/acentos-insensitive).
    """
    # garante pasta de saída
    os.makedirs(output_dir, exist_ok=True)

    creds = _get_credentials(client_secrets_path=client_secrets_path, token_path=token_path)
    service = build("sheets", "v4", credentials=creds)

    try:
        values = _read_sheet_values(service, spreadsheet_id, sheet_name, header_row=10)
        if not values:
            raise RuntimeError("Nenhum dado retornado. Verifique a aba e o ID da planilha.")

        headers, rows = values[0], values[1:]
        idx_remessa = _find_col_idx(headers, REMESSA_DATE_KEYS, fallback_idx=13)  # N = 13 (0-based)
        idx_recurso = _find_col_idx(headers, RECURSO_DATE_KEYS, fallback_idx=26)  # AA = 26 (0-based)
        idx_valor_recursado = _find_col_idx(headers, VALOR_RECURSADO_KEYS, fallback_idx=None)

        headers_map = _build_header_map(headers)
        out_rows: List[dict] = []

        for row in rows:
            if all(str(c).strip() == "" for c in row):
                continue

            # dict {header_original: valor}
            row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}

            d_remessa = _as_date(row[idx_remessa]) if (idx_remessa is not None and len(row) > idx_remessa) else None
            d_recurso = _as_date(row[idx_recurso]) if (idx_recurso is not None and len(row) > idx_recurso) else None

            if not ((d_remessa == data_pagamento) or (d_recurso == data_pagamento)):
                continue

            base_item = _extract_required_fields(row_dict, headers_map)

            # Valor recursado com prioridade para o cabeçalho exato
            valor_recursado = None
            if idx_valor_recursado is not None and len(row) > idx_valor_recursado:
                valor_recursado = row[idx_valor_recursado]
            if valor_recursado in (None, ""):
                for k in VALOR_RECURSADO_KEYS:
                    exact = next((h for h in headers if _norm(h) == _norm(k)), None)
                    if exact and (row_dict.get(exact) not in (None, "")):
                        valor_recursado = row_dict.get(exact)
                        break

            base_item["Valor recursado"] = valor_recursado

            base_item["Data Remessa"] = d_remessa.isoformat() if d_remessa else None
            base_item["Data Recurso"] = d_recurso.isoformat() if d_recurso else None
            base_item["_match_remessa"] = (d_remessa == data_pagamento)
            base_item["_match_recurso"] = (d_recurso == data_pagamento)

            out_rows.append(base_item)

        # grava JSON
        ymd = data_pagamento.strftime("%Y%m%d")
        safe_sheet = "".join(c for c in sheet_name if c.isalnum() or c in ("_", "-")).strip() or "aba"
        fpath = os.path.join(output_dir, f"saida_{safe_sheet}_{ymd}.json")
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(out_rows, f, ensure_ascii=False, indent=2)

        return fpath

    except HttpError as e:
        raise RuntimeError(f"Erro ao acessar Google Sheets: {e}") from e
    
# --------------------------------------------------------------------------- #
# Função: processar_nfse_para_json
# --------------------------------------------------------------------------- #
def _normalize_nf_number(x):
    s = "" if x is None else str(x)
    return "".join(ch for ch in s if ch.isdigit())

def processar_nfse_para_json(
    spreadsheet_id: str,
    sheet_name: str,
    modo: str,                 # "data" ou "numero"
    valor: str,                # data ISO (YYYY-MM-DD) quando modo="data", ou número da NF quando modo="numero"
    output_dir: str = ".",
    client_secrets_path: str = "client_secret.json",
    token_path: str = "token.json",
) -> str:
    """Gera JSON de NFS-e por *data de emissão* (“Data envio NF - Convenio”)
    ou por *número da NF* (“Nº NF” / “NF recurso”). Marca linhas que casam em
    NF recurso para o front-end aplicar regras visuais."""
    import os, json
    from datetime import date
    os.makedirs(output_dir, exist_ok=True)

    creds = _get_credentials(client_secrets_path=client_secrets_path, token_path=token_path)
    service = build("sheets", "v4", credentials=creds)

    values = _read_sheet_values(service, spreadsheet_id, sheet_name, header_row=10)
    if not values:
        raise RuntimeError("Nenhum dado retornado. Verifique a aba e o ID da planilha.")

    headers, rows = values[0], values[1:]
    headers_map = _build_header_map(headers)

    # índices dos campos necessários
    idx_emissao = _find_col_idx(headers, ["data envio nf - convenio"], fallback_idx=None)

    header_num_nf = next((h for h in headers if _norm(h) in ("nº nf","n nf","no nf","numero nf","nfse")), None)
    header_nf_recurso = next((h for h in headers if _norm(h) in ("nf recurso","nf de recurso")), None)

    # onde procurar "Valor recursado"
    idx_valor_recursado = _find_col_idx(headers, VALOR_RECURSADO_KEYS, fallback_idx=None)

    out_rows = []
    if modo == "data":
        target_date = date.fromisoformat(str(valor))

    def _get_valor_recursado(row_dict, row):
        """Resolve 'Valor recursado' com prioridade para cabeçalhos equivalentes."""
        # 1) se tem índice direto
        if idx_valor_recursado is not None and len(row) > idx_valor_recursado:
            v = row[idx_valor_recursado]
            if str(v).strip() != "":
                return v
        # 2) varrer candidatos por nome normalizado
        for k in VALOR_RECURSADO_KEYS:
            exact = next((h for h in headers if _norm(h) == _norm(k)), None)
            if exact and (row_dict.get(exact) not in (None, "")):
                return row_dict.get(exact)
        return None

    for row in rows:
        if all(str(c).strip() == "" for c in row):
            continue

        row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        base_item = _extract_required_fields(row_dict, headers_map)

        # captura data de emissão da NF
        d_emissao = _as_date(row[idx_emissao]) if idx_emissao is not None and len(row) > idx_emissao else None

        # por padrão (preenchidos embaixo)
        base_item["_nfse_modo"] = modo
        base_item["_nfse_match_kind"] = None

        match_emissao = match_nf = False

        if modo == "data":
            match_emissao = (d_emissao == target_date)
            if not match_emissao:
                continue
        else:
            # normaliza números p/ comparação
            target_nf = _normalize_nf_number(valor)
            cand_nf = _normalize_nf_number(row_dict.get(header_num_nf)) if header_num_nf else ""
            cand_nfrec = _normalize_nf_number(row_dict.get(header_nf_recurso)) if header_nf_recurso else ""
            if not (target_nf and (target_nf == cand_nf or target_nf == cand_nfrec)):
                continue
            match_nf = True
            # classifica o tipo de match
            if target_nf == cand_nfrec and cand_nfrec != "":
                base_item["_nfse_match_kind"] = "recurso"
            else:
                base_item["_nfse_match_kind"] = "nf"

        # adiciona valor recursado ao JSON (útil para a troca no front)
        base_item["Valor recursado"] = _get_valor_recursado(row_dict, row)

        base_item["Data Emissao NF"] = d_emissao.isoformat() if d_emissao else None
        base_item["_match_emissao"] = match_emissao
        base_item["_match_nf"] = match_nf

        out_rows.append(base_item)

    safe_sheet = "".join(c for c in sheet_name if c.isalnum() or c in ("_", "-")).strip() or "aba"
    suffix = f"data_{target_date.strftime('%Y%m%d')}" if modo == "data" else f"nf_{_normalize_nf_number(valor) or 'sem_numero'}"
    fname = f"saida_nfse_{safe_sheet}_{suffix}.json"
    fpath = os.path.join(output_dir, fname)

    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(out_rows, f, ensure_ascii=False, indent=2)
    return fpath



# ------------------------- Capa (NFSe emitidas no dia) -----------------------

def gerar_capa_nfse_por_data(
    spreadsheet_id: str,
    sheet_names: list[str],
    data_emissao: str,          # ISO YYYY-MM-DD
    output_dir: str = ".",
    client_secrets_path: str = "client_secret.json",
    token_path: str = "token.json",
) -> str:
    """Gera o JSON da aba **Capa**.

    Varre todas as abas/convênios em `sheet_names`, filtra pelo dia em
    “Data envio NF - Convenio” e consolida por (NFSe, Convênio).

    Regra do valor quando a NFSe aparece em mais de uma remessa:
      - `valor_nf_base` = MAIOR "Valor NF" encontrado para a mesma NFSe
      - `valor_recursado_total` = SOMA de "Valor recursado" (quando existir)
      - `valor_total` = valor_nf_base + valor_recursado_total

    Saída: lista de dicts: { "NFSe": str, "Convenio": str, "Valor": float }
    """
    import os, json, re
    from datetime import date
    import pandas as pd

    os.makedirs(output_dir, exist_ok=True)

    target_date = date.fromisoformat(str(data_emissao))

    creds = _get_credentials(client_secrets_path=client_secrets_path, token_path=token_path)
    service = build("sheets", "v4", credentials=creds)

    def _money_to_float(v) -> float:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        try:
            return float(v)
        except Exception:
            s = str(v).strip()
            if s == "":
                return 0.0
            s = s.replace("R$", "").replace(" ", "").replace("\u00A0", "")
            s = s.replace(".", "").replace(",", ".")
            try:
                return float(s)
            except Exception:
                return 0.0

    # acumuladores por (nfse, convenio): max(valor_nf) e soma(valor_recursado)
    acc: dict[tuple[str, str], dict[str, float]] = {}
    nfse_do_dia: set[tuple[str, str]] = set()  # (nfse, sheet_name)

    for sheet_name in (sheet_names or []):
        values = _read_sheet_values(service, spreadsheet_id, sheet_name, header_row=10)
        if not values:
            continue

        headers, rows = values[0], values[1:]

        idx_emissao = _find_col_idx(headers, ["data envio nf - convenio"], fallback_idx=None)
        if idx_emissao is None:
            continue

        idx_num_nf = _find_col_idx(
            headers,
            ["nº nf", "n nf", "no nf", "numero nf", "nfse", "nfs-e", "nfs e"],
            fallback_idx=None,
        )
        if idx_num_nf is None:
            continue

        idx_valor_nf = _find_col_idx(
            headers,
            ["valor nf", "valor nfse", "valor nfs-e", "valor da nf"],
            fallback_idx=None,
        )
        idx_valor_recursado = _find_col_idx(headers, VALOR_RECURSADO_KEYS, fallback_idx=None)

        idx_nf_recurso = _find_col_idx(
            headers,
            ["nf recurso", "nf_rec", "nf rec", "nf de recurso", "nf recurso glosa"],
            fallback_idx=None,
        )
        idx_data_pgto_recurso = _find_col_idx(headers, RECURSO_DATE_KEYS, fallback_idx=None)

        # -------------------------
        # 1ª PASSAGEM: NFSe emitidas no dia (remessas "normais")
        # -------------------------
        for row in rows:
            if idx_emissao >= len(row) or idx_num_nf >= len(row):
                continue

            d = _as_date(row[idx_emissao])
            if not d or d != target_date:
                continue

            nfse = _normalize_nf_number(row[idx_num_nf])
            if not nfse:
                continue

            # marca que essa NFSe pertence ao dia (por convênio/aba)
            nfse_do_dia.add((str(nfse), str(sheet_name)))

        # -------------------------
        # 1B) Marca NF alvo por "Data pgto recurso" (NF recurso pago no dia)
        # -------------------------
        if idx_nf_recurso is not None and idx_data_pgto_recurso is not None:
            for row in rows:
                if idx_nf_recurso >= len(row) or idx_data_pgto_recurso >= len(row):
                    continue

                dpg = _as_date(row[idx_data_pgto_recurso])
                if not dpg or dpg != target_date:
                    continue

                nf_recurso = _normalize_nf_number(row[idx_nf_recurso])
                if not nf_recurso:
                    continue

                nfse_do_dia.add((str(nf_recurso), str(sheet_name)))
        
        # -------------------------
        # 1C) Soma Valor NF de TODAS as linhas cuja Nº NF esteja no conjunto alvo
        #      (assim NF que entrou no set alvo também traz todas as remessas)
        # -------------------------
        for row in rows:
            if idx_num_nf >= len(row):
                continue

            nfse = _normalize_nf_number(row[idx_num_nf])
            if not nfse:
                continue

            key = (str(nfse), str(sheet_name))
            if key not in nfse_do_dia:
                continue

            valor_nf = 0.0
            if idx_valor_nf is not None and idx_valor_nf < len(row):
                valor_nf = _money_to_float(row[idx_valor_nf])

            valor_rec = 0.0
            if idx_valor_recursado is not None and idx_valor_recursado < len(row):
                valor_rec = _money_to_float(row[idx_valor_recursado])

            slot = acc.get(key)
            if slot is None:
                acc[key] = {"sum_valor_nf": float(valor_nf), "sum_valor_rec": float(valor_rec)}
            else:
                slot["sum_valor_nf"] += float(valor_nf)
                slot["sum_valor_rec"] += float(valor_rec)

                
        # -------------------------
        # 2ª PASSAGEM: Recurso de Glosa (NF recurso) SEM filtrar pela data de emissão
        # -------------------------
        if idx_nf_recurso is not None:
            for row in rows:
                if idx_nf_recurso >= len(row):
                    continue

                nf_recurso = _normalize_nf_number(row[idx_nf_recurso])
                if not nf_recurso:
                    continue

                # só soma se o NF recurso apontar para uma NFSe que foi emitida no dia
                key_rec = (str(nf_recurso), str(sheet_name))
                if key_rec not in nfse_do_dia:
                    continue

                valor_rec = 0.0
                if idx_valor_recursado is not None and idx_valor_recursado < len(row):
                    valor_rec = _money_to_float(row[idx_valor_recursado])

                slot2 = acc.get(key_rec)
                if slot2 is None:
                    acc[key_rec] = {"sum_valor_nf": 0.0, "sum_valor_rec": float(valor_rec)}
                else:
                    slot2["sum_valor_rec"] += float(valor_rec)

    out_rows = []
    def _nf_sort(n: str) -> int:
        return int(re.sub(r"\D", "", n) or "0")

    for (nfse, convenio), v in sorted(acc.items(), key=lambda x: (x[0][1], _nf_sort(x[0][0]))):
        total = float(v.get("sum_valor_nf", 0.0)) + float(v.get("sum_valor_rec", 0.0)) 
        out_rows.append({"NFSe": str(nfse), "Convenio": str(convenio), "Valor": round(total, 2)})

    fname = f"saida_capa_nfse_{target_date.strftime('%Y%m%d')}.json"
    fpath = os.path.join(output_dir, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(out_rows, f, ensure_ascii=False, indent=2)
    return fpath

# ------------------------- DataFrames e totais ------------------------------

def make_remessas_df(items: List[Dict[str, Any]], selected_date: date) -> pd.DataFrame:
    sel_iso = selected_date.isoformat() if selected_date else None
    base = []
    for x in items:
        d_rem = x.get("Data Remessa"); vpgto = _as_number(x.get("Valor pgto"))
        if d_rem == sel_iso and vpgto is not None and vpgto != 0: base.append(x)
    df = pd.DataFrame(base)
    for col in REMESSAS_COLUMNS:
        if col not in df.columns: df[col] = None
    df = df[REMESSAS_COLUMNS]
    for c in ["Valor envio XML - Remessa","Valor pgto","Valor glosado","Imposto","Glosa mantida"]:
        df[c] = df[c].apply(_as_number)
    return df

def make_recursos_df(items: List[Dict[str, Any]], data_pagamento: date) -> pd.DataFrame:
    registros = []
    alvo_iso = data_pagamento.isoformat()

    for it in items:
        if it.get("Data Recurso") != alvo_iso:
            continue

        registros.append({
            "Nº Remessa": it.get("Nº Remessa"),
            "Ref.": it.get("Ref."),
            "Nº NF": it.get("NF recurso") or it.get("Nº NF"),  # exibe sempre uma NF
            "Valor recursado": it.get("Valor recursado"),
            "Valor pago": it.get("Valor pago"),
            "Imposto": it.get("Imposto"),
            "Glosa mantida": it.get("Glosa mantida"),
        })

    if not registros:
        return pd.DataFrame(columns=[
            "Nº Remessa","Ref.","Nº NF",
            "Valor recursado","Valor pago",
            "Imposto","Glosa mantida",
        ])

    df = pd.DataFrame(registros)

    for col in ["Valor recursado","Valor pago","Imposto","Glosa mantida"]:
        if col in df.columns:
            df[col] = df[col].apply(_as_number)

    return df

def sum_col(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns: return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())

def compute_totals_remessas(df: pd.DataFrame) -> Dict[str, float]:
    return {
        "total_envio_xml": sum_col(df,"Valor envio XML - Remessa"),
        "total_pgto": sum_col(df,"Valor pgto"),
        "total_glosado": sum_col(df,"Valor glosado"),
        "total_imposto": sum_col(df,"Imposto"),
        "total_glosa_mantida": sum_col(df,"Glosa mantida"),
        "total_pago": sum_col(df,"Valor pgto"),
    }

def compute_totals_recursos(df: pd.DataFrame) -> Dict[str, float]:
    return {
        "total_glosa": sum_col(df,"Valor glosado") if "Valor glosado" in df.columns else 0.0,
        "total_imposto": sum_col(df,"Imposto"),
        "total_glosa_mantida": sum_col(df,"Glosa mantida"),
        "total_pago": sum_col(df,"Valor pago") if "Valor pago" in df.columns else sum_col(df,"Valor pgto"),
    }

# ------------------------- Helpers CSV e formatação -------------------------

def _ensure_len(row, n=CSV_COLS):
    row = list(row)
    if len(row) < n: row += [""] * (n - len(row))
    elif len(row) > n: row = row[:n]
    return row

def _fmt_ref_mmYYYY(ref_raw):
    import unicodedata
    s_raw = str(ref_raw or "").strip()
    if not s_raw: return ""
    s = unicodedata.normalize("NFD", s_raw).lower()
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace(".","").replace("_"," ").strip()
    m = re.search(r"(\d{1,2})[\/\-\s\.](\d{2,4})", s)
    if m:
        mm = int(m.group(1)); yy = m.group(2); yyyy = int(yy) + 2000 if len(yy)==2 else int(yy)
        if 1 <= mm <= 12: return f"{mm:02d}/{yyyy}"
    month_map = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,"jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12,
                 "feb":2,"apr":4,"may":5,"aug":8,"sep":9,"oct":10,"dec":12}
    m = re.search(r"([a-z]{3,4})[\/\-\s]+(\d{2,4})", s)
    if m:
        mon = m.group(1)[:3]; yy = m.group(2)
        if mon in month_map:
            mm = month_map[mon]; yyyy = int(yy) + 2000 if len(yy)==2 else int(yy)
            return f"{mm:02d}/{yyyy}"
    m = re.search(r"(\d{4})[\/\-\s\.](\d{1,2})", s)
    if m:
        yyyy = int(m.group(1)); mm = int(m.group(2))
        if 1 <= mm <= 12: return f"{mm:02d}/{yyyy}"
    return s_raw

def _fmt_amount_csv(v):
    if v is None: return "0"
    try: fv = float(v)
    except Exception:
        s = str(v).strip().replace("R$","").replace(" ","")
        if "," in s and "." in s: s = s.replace(".","").replace(",",".")
        elif "," in s: s = s.replace(",",".")
        try: fv = float(s)
        except Exception: return "0"
    cents = int(round(fv * 100))
    return str(cents)

def _slugify(s: str) -> str:
    import unicodedata as _ud, re as _re
    s = str(s or "")
    s = _ud.normalize("NFKD", s)
    s = "".join(c for c in s if not _ud.combining(c))
    s = s.lower().strip()
    s = _re.sub(r"[^a-z0-9]+","_",s)
    s = _re.sub(r"_+","_",s).strip("_")
    return s


def gerar_csv_glosa_mantida_bytes(
    items: list[dict],
    encoding: str = "cp1252",
):
    """Gera CSV de lançamentos contábeis para *Glosa Mantida* (manual).

    Cada item gera 2 linhas (padrão do seu template de importação):
      - Débito:  3149021 / restrição "2"  / valor positivo
      - Crédito: 1133004 / restrição "0A" / valor negativo

    Observações:
      - A subconta é a mesma lógica dos demais relatórios (SUBCONTA_MAP + overrides).
      - A 1ª linha do CSV ("header_first_row") segue o mesmo padrão dos outros CSVs.
    """

    import io
    import csv as _csv
    import unicodedata

    # ---- prefs / overrides / unidade (igual outras funções de CSV) ----
    prefs: dict = {}
    ovrs: dict = {}
    unidade: str = ""
    if st is not None:
        try:
            prefs = st.session_state.get("csv_prefs", {}) or {}
            ovrs = st.session_state.get("csv_convenio_overrides", {}) or {}
            unidade = st.session_state.get("unidade", "") or ""
        except Exception:
            prefs, ovrs, unidade = {}, {}, ""

    header_first_row = prefs.get("header_first_row")
    if not header_first_row:
        header_first_row = (
            ("1941", "Clínica Adventista de Curitiba - IASBS")
            if (unidade or "").upper() == "CMAC"
            else ("1841", "Clínica Adventista de Porto Alegre - IASBS")
        )

    # Mesmo SUBCONTA_MAP usado nas outras funções
    SUBCONTA_MAP = {
        "Afpergs": "13",
        "Proasa": "51",
        "Amil": "39",
        "Assefaz": "27",
        "Banco Central": "43",
        "Cassi": "26",
        "Doctor": "54",
        "Embratel": "18",
        "Humana": "55",
        "GEAP": "19",
        "Geap": "19",
        "Postal Saúde": "29",
        "Prevent Senior": "91",
        "Saúde Caixa": "15",
        "Medservice": "16",
        "Life": "45",
        "Gente Saúde": "59",
        "Capesesp": "20",
        "Ipê Saúde": "12",
    }

    def _nfc(s):
        return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

    rows = [_ensure_len(list(header_first_row), CSV_COLS)]

    total = 0.0
    for it in (items or []):
        try:
            convenio = str(it.get("convenio") or "").strip()
            desc = str(it.get("descricao") or it.get("desc") or "").strip()
            val = it.get("valor")

            # converte valor
            try:
                val_f = float(val)
            except Exception:
                s = str(val).strip().replace("R$", "").replace(" ", "")
                if "," in s and "." in s:
                    s = s.replace(".", "").replace(",", ".")
                elif "," in s:
                    s = s.replace(",", ".")
                val_f = float(s) if s else 0.0

            if not convenio or not desc or val_f == 0:
                continue

            oconv: dict = {}
            if isinstance(ovrs, dict):
                oconv = ovrs.get(convenio, {}) or {}

            subconta_convenio = (
                oconv.get("subconta_convenio")
                or SUBCONTA_MAP.get(convenio, "13")
            )

            total += float(val_f)

            # Débito
            rows.append(_ensure_len([
                "3149021", subconta_convenio, "10", "1110", "2",
                _fmt_amount_csv(val_f), "N", _nfc(desc),
            ], CSV_COLS))

            # Crédito
            rows.append(_ensure_len([
                "1133004", subconta_convenio, "10", "1110", "0A",
                "-" + _fmt_amount_csv(val_f), "N", _nfc(desc),
            ], CSV_COLS))

        except Exception:
            continue

    sio = io.StringIO()
    writer = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
    writer.writerows(rows)
    csv_text = sio.getvalue()
    csv_bytes = csv_text.encode(encoding, errors="replace")

    from datetime import date as _date

    date_str = _date.today().strftime("%Y%m%d")
    n = int(len(items or []))
    fname = f"lanc_glosa_mantida_{date_str}_n{n}_total{float(total):.2f}.csv"
    return fname, csv_bytes

# ------------------------- Geração do CSV de lançamentos --------------------
# (mantido igual ao original do usuário, apenas limpeza leve e comentários)
# ... (mantido integralmente a partir daqui) ...

def gerar_csv_lancamentos_bytes(
    df_remessas: pd.DataFrame,
    rem_totals: Dict[str, float],
    df_recursos: pd.DataFrame,
    rec_totals: Dict[str, float],
    selected_date: date,
    selected_convenio: str,
    modelo_csv_path: Optional[str] = None,
    encoding: str = "cp1252",
):
    import io, csv, unicodedata

    prefs: dict = {}
    ovrs: dict = {}
    unidade: str = ""
    if st is not None:
        try:
            prefs = st.session_state.get("csv_prefs", {}) or {}
            ovrs = st.session_state.get("csv_convenio_overrides", {}) or {}
            unidade = st.session_state.get("unidade", "") or ""
        except Exception:
            prefs, ovrs, unidade = {}, {}, ""

    oconv: dict = {}
    if isinstance(ovrs, dict):
        oconv = ovrs.get(selected_convenio, {}) or {}

    header_first_row = prefs.get("header_first_row")
    if not header_first_row:
        if (unidade or "").upper() == "CMAC":
            header_first_row = ("1941", "Clínica Adventista de Curitiba - IASBS")
        else:
            header_first_row = ("1841", "Clínica Adventista de Porto Alegre - IASBS")

    deposito_subconta_banco = (
        oconv.get("deposito_subconta_banco")
        or prefs.get("deposito_subconta_banco")
        or ("03645/00044210" if (unidade or "").upper() == "CMAC" else "3708/0001649-7")
    )

    deposito_prefix = (prefs.get("deposito_prefix") or ("Ac. Depósito")).strip()
    if not deposito_prefix.lower().startswith("ac."):
        deposito_prefix = f"Ac. {deposito_prefix.lstrip()}"

    oconv = ovrs.get(selected_convenio, {})

    SUBCONTA_MAP = {
        "Afpergs": "13", "Proasa": "51", 
        "Amil": "39","Assefaz": "27","Banco Central": "43","Cassi": "26","Doctor": "54",
        "Embratel": "18","Humana": "55","GEAP": "19","Geap": "19",
        "Postal Saúde": "29","Prevent Senior": "91","Saúde Caixa": "15", "Mediservice": "16", "Life": "45",
        "Gente Saúde": "59", "Ipê Saúde": "12",
    }
    DEPOSITO_SUFFIX = {
        "Afpergs": "Assoc Func Publicos", "Proasa": "Proasa Programa Adven",
        "Amil": "Amil Assistencia Med","Assefaz": "Fundacao Assistencia",
        "Banco Central": "Pagamento Sigef Ap","Cassi": "Caixa de Assistencia",
        "Doctor": "Doctor Clin Ope De P","Embratel": "Telos Fundacao Embrat",
        "Humana": "Humana Saude Ltda","GEAP": "Geap Autogestao Em S","Geap": "Geap Autogestao Em S",
        "Postal Saúde": "Postal Saúde","Prevent Senior": "Prevent Senior Privat","Saúde Caixa": "Saúde Caixa",
        "Mediservice": "Ac. Mediservice Operadora Planos Sau","Life": "Ac. Life Empresarial Saúde Ltda",
        "Gente Saúde": "Ac. Gente Clube - Benefi", "Ipê Saúde": "Ac. Ipê Saúde",
    }

    subconta_convenio = oconv.get("subconta_convenio") or SUBCONTA_MAP.get(selected_convenio, "13")
    deposito_suffix = oconv.get("deposito_suffix") or DEPOSITO_SUFFIX.get(selected_convenio, selected_convenio)
    glosa_neg_target = oconv.get("glosa_neg_target") or ("1133003" if selected_convenio in ("Afpergs", "Proasa") else "1133001")

    def _nfc(s):
        return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

    def _norm_str(x):
        return (str(x).strip() if x is not None else "").replace("\u00A0", " ")

    def _to_float(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        try:
            return float(v)
        except Exception:
            s = str(v).strip().replace("R$", "").replace(" ", "")
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            try:
                return float(s)
            except Exception:
                return None

    def _dedupe(df, keys):
        if df is None or len(df) == 0:
            return df
        df = df.copy()
        for c in ["Nº Remessa","Nº NF","Ref."]:
            if c in df.columns:
                df[c] = df[c].apply(_norm_str)
        for c in keys:
            if c in df.columns:
                df[c] = df[c].apply(_to_float)
        subset = [c for c in ["Nº Remessa","Nº NF","Ref."] + keys if c in df.columns]
        if subset:
            df = df.drop_duplicates(subset=subset, keep="first")
        else:
            df = df.drop_duplicates(keep="first")
        return df.dropna(how="all")

    df_remessas = _dedupe(df_remessas, ["Valor pgto","Valor glosado"])
    df_recursos = _dedupe(df_recursos, ["Valor pago"])

    rows = [_ensure_len(list(header_first_row), CSV_COLS)]

    total_debito = (rem_totals.get("total_pago") or 0.0) + (rec_totals.get("total_pago") or 0.0)

    # --- monta descrição do depósito SEMPRE com data; se o sufixo já vier com "Ac.", removemos para não duplicar
    dep_suffix_clean = str(deposito_suffix or "").strip()
    if dep_suffix_clean.lower().startswith("ac."):
        dep_suffix_clean = dep_suffix_clean[3:].lstrip(" -–—").strip()

    desc_deposito = _nfc(f"{deposito_prefix} {selected_date.strftime('%d/%m/%Y')} - {dep_suffix_clean}")


    rows.append(_ensure_len([
        "2162001", deposito_subconta_banco, "10", "1110", "0A",
        _fmt_amount_csv(total_debito), "N", desc_deposito
    ], CSV_COLS))

    deposito_prefix = (prefs.get("deposito_prefix") or "Ac. Depósito").strip()
    if not deposito_prefix.lower().startswith("ac."):
        deposito_prefix = f"Ac. {deposito_prefix.lstrip()}"

    if df_remessas is not None and len(df_remessas) > 0:
        for _, r in df_remessas.iterrows():
            num_remessa = str(r.get("Nº Remessa") or "").strip()
            num_nf = str(r.get("Nº NF") or "").strip()
            ref_fmt = _fmt_ref_mmYYYY(str(r.get("Ref.") or "").strip())

            val_pgto = r.get("Valor pgto")
            try:
                val_pgto = float(val_pgto) if val_pgto is not None else None
            except Exception:
                val_pgto = None
            if val_pgto is not None and not pd.isna(val_pgto) and val_pgto != 0:
                desc_rem = _nfc(f"Rec. Rem. {num_remessa} - NFSe {num_nf} - Fat. {ref_fmt}")
                rows.append(_ensure_len([
                    "1133001", subconta_convenio, "10", "1110", "0A",
                    "-" + _fmt_amount_csv(val_pgto), "N", desc_rem
                ], CSV_COLS))

            val_glosa = r.get("Valor glosado") or 0
            try:
                val_glosa = float(val_glosa)
            except Exception:
                val_glosa = 0.0

            if abs(val_glosa) > 0:
                if glosa_neg_target == "1133001":
                    desc_g = _nfc(f"Glosa S/Rem. {num_remessa} - NFSe {num_nf} - Fat. {ref_fmt}")
                else:
                    desc_g = _nfc(f"Glosa S/Rem. {num_remessa} - Faturamento Plano de Saúde - {ref_fmt}")

                rows.append(_ensure_len([
                    "1133004", subconta_convenio, "10", "1110", "0A",
                    _fmt_amount_csv(val_glosa), "N", desc_g
                ], CSV_COLS))
                rows.append(_ensure_len([
                    glosa_neg_target, subconta_convenio, "10", "1110", "0A",
                    "-" + _fmt_amount_csv(val_glosa), "N", desc_g
                ], CSV_COLS))

    if df_recursos is not None and len(df_recursos) > 0:
        for _, r in df_recursos.iterrows():
            num_remessa = str(r.get("Nº Remessa") or "").strip()
            num_nf = str(r.get("Nº NF") or "").strip()
            ref_fmt = _fmt_ref_mmYYYY(str(r.get("Ref.") or "").strip())

            val_pago = r.get("Valor pago")
            try:
                val_pago = float(val_pago) if val_pago is not None else None
            except Exception:
                val_pago = None
            if val_pago is not None and not pd.isna(val_pago) and val_pago != 0:
                desc_rec = _nfc(f"Rec. Rem. {num_remessa} - NFSe {num_nf} - Rg. {ref_fmt}")
                rows.append(_ensure_len([
                    "1133001", subconta_convenio, "10", "1110", "0A",
                    "-" + _fmt_amount_csv(val_pago), "N", desc_rec
                ], CSV_COLS))

    import io, csv as _csv
    sio = io.StringIO()
    writer = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
    writer.writerows(rows)
    csv_text = sio.getvalue()
    csv_bytes = csv_text.encode(encoding, errors="replace")

    conv_slug = _slugify(selected_convenio)
    date_str = selected_date.strftime("%Y%m%d")
    n_rem = int(len(df_remessas) if df_remessas is not None else 0)
    n_rec = int(len(df_recursos) if df_recursos is not None else 0)
    total_str = f"{float(total_debito):.2f}"
    fname = f"lanc_{conv_slug}_{date_str}_rem{n_rem}_rec{n_rec}_total{total_str}.csv"
    return fname, csv_bytes

def gerar_csv_recursos_bytes(
    df_recursos: pd.DataFrame,
    rec_totals: Dict[str, float],
    df_remessas: Optional[pd.DataFrame],
    rem_totals: Dict[str, float],
    selected_date: date,
    selected_convenio: str,
    modelo_csv_path: Optional[str] = None,
    encoding: str = "cp1252",
    add_debito_if_only_resources: bool = True,
):
    import io, csv as _csv, unicodedata

    prefs: dict = {}
    ovrs: dict = {}
    unidade: str = ""
    if st is not None:
        try:
            prefs = st.session_state.get("csv_prefs", {}) or {}
            ovrs = st.session_state.get("csv_convenio_overrides", {}) or {}
            unidade = st.session_state.get("unidade", "") or ""
        except Exception:
            prefs, ovrs, unidade = {}, {}, ""

    oconv: dict = {}
    if isinstance(ovrs, dict):
        oconv = ovrs.get(selected_convenio, {}) or {}

    header_first_row = prefs.get("header_first_row")
    if not header_first_row:
        header_first_row = ("1941", "Clínica Adventista de Curitiba - IASBS") if (unidade or "").upper() == "CMAC" \
                           else ("1841", "Clínica Adventista de Porto Alegre - IASBS")

    deposito_subconta_banco = (
        oconv.get("deposito_subconta_banco")
        or prefs.get("deposito_subconta_banco")
        or ("03645/00044210" if (unidade or "").upper() == "CMAC" else "3708/0001649-7")
    )

    deposito_prefix = (prefs.get("deposito_prefix") or "Ac. Depósito").strip()
    if not deposito_prefix.lower().startswith("ac."):
        deposito_prefix = f"Ac. {deposito_prefix.lstrip()}"

    SUBCONTA_MAP = {
        "Afpergs": "13", "Proasa": "51",
        "Amil": "39","Assefaz": "27","Banco Central": "43","Cassi": "26","Doctor": "54",
        "Embratel": "18","Humana": "55","GEAP": "19","Geap": "19",
        "Postal Saúde": "29","Prevent Senior": "91","Saúde Caixa": "15",
        "Medservice": "16", "Life": "45", "Gente Saúde": "59", "Capesesp": "20", "Ipê Saúde": "12",
    }
    DEPOSITO_SUFFIX = {
        "Afpergs": "Assoc Func Publicos", "Proasa": "Proasa",
        "Amil": "Amil Assistencia Med","Assefaz": "Fundacao Assistencia",
        "Banco Central": "Pagamento Sigef Ap","Cassi": "Caixa de Assistencia",
        "Doctor": "Doctor Clin Ope De P","Embratel": "Telos Fundacao Embrat",
        "Humana": "Humana Saude Ltda","GEAP": "Geap Autogestao Em S","Geap": "Geap Autogestao Em S",
        "Postal Saúde": "Postal Saúde","Prevent Senior": "Prevent Senior Privat","Saúde Caixa": "Saúde Caixa",
        "Mediservice": "Ac. Mediservice Operadora Planos Sau",
        "Life": "Ac. Life Empresarial Saúde Ltda",
        "Gente Saúde": "Ac. Gente Clube - Benefi", "Capesesp": "Caixa De Previdencia", "Ipê Saúde": "Ipê Saúde",
    }

    subconta_convenio = oconv.get("subconta_convenio") or SUBCONTA_MAP.get(selected_convenio, "13")
    deposito_suffix = oconv.get("deposito_suffix") or DEPOSITO_SUFFIX.get(selected_convenio, selected_convenio)
    glosa_neg_target = oconv.get("glosa_neg_target") or ("1133003" if selected_convenio in ("Afpergs", "Proasa") else "1133001")

    def _nfc(s): return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

    def _norm_str(x): return (str(x).strip() if x is not None else "").replace("\u00A0", " ")
    def _to_float(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return None
        try: return float(v)
        except Exception:
            s = str(v).strip().replace("R$","").replace(" ","")
            if "," in s and "." in s: s = s.replace(".","").replace(",",".")
            elif "," in s: s = s.replace(",",".")
            try: return float(s)
            except Exception: return None
    def _dedupe_recursos(df):
        if df is None or len(df) == 0: return df
        df = df.copy()
        for c in ["Nº Remessa","Nº NF","Ref."]:
            if c in df.columns: df[c] = df[c].apply(_norm_str)
        if "Valor pago" in df.columns: df["Valor pago"] = df["Valor pago"].apply(_to_float)
        subset = [c for c in ["Nº Remessa","Nº NF","Ref.","Valor pago"] if c in df.columns]
        if subset: df = df.drop_duplicates(subset=subset, keep="first")
        else: df = df.drop_duplicates(keep="first")
        return df.dropna(how="all")

    df_recursos = _dedupe_recursos(df_recursos)

    rows = [_ensure_len(list(header_first_row), CSV_COLS)]

    total_remessas_pago = float(rem_totals.get("total_pago") or 0.0)
    total_recursos_pago = float(rec_totals.get("total_pago") or 0.0)
    is_only_resources_day = (total_remessas_pago == 0.0) and (df_remessas is None or len(df_remessas) == 0)

    if add_debito_if_only_resources and is_only_resources_day and total_recursos_pago > 0:
        dep_suffix_clean = str(deposito_suffix or "").strip()
        if dep_suffix_clean.lower().startswith("ac."):
            dep_suffix_clean = dep_suffix_clean[3:].lstrip(" -–—").strip()

        desc_deposito = _nfc(f"{deposito_prefix} {selected_date.strftime('%d/%m/%Y')} - {dep_suffix_clean}")

        rows.append(_ensure_len([
            "2162001", deposito_subconta_banco, "10", "1110", "0A",
            _fmt_amount_csv(total_recursos_pago), "N", desc_deposito
        ], CSV_COLS))

    if df_recursos is not None and len(df_recursos) > 0:
        for _, r in df_recursos.iterrows():
            num_remessa = str(r.get("Nº Remessa") or "").strip()
            num_nf = str(r.get("Nº NF") or "").strip()
            ref_fmt = _fmt_ref_mmYYYY(str(r.get("Ref.") or "").strip())

            val_pago = r.get("Valor pago")
            try:
                val_pago = float(val_pago) if val_pago is not None else None
            except Exception:
                val_pago = None
            if val_pago is None or pd.isna(val_pago) or val_pago == 0:
                continue

            if glosa_neg_target == "1133001":
                conta = "1133004"
                desc_rec = _nfc(f"Rec. Glosa S/Rem. {num_remessa} - NFSe {num_nf} - Fat. {ref_fmt}")
            else:
                conta = "1133001"
                desc_rec = _nfc(f"Rec. Rem. {num_remessa} - NFSe {num_nf} - Rg. {ref_fmt}")

            rows.append(_ensure_len([
                conta, subconta_convenio, "10", "1110", "0A",
                "-" + _fmt_amount_csv(val_pago), "N", desc_rec
            ], CSV_COLS))

    sio = io.StringIO()
    writer = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
    writer.writerows(rows)
    csv_text = sio.getvalue()
    csv_bytes = csv_text.encode(encoding, errors="replace")

    conv_slug = _slugify(selected_convenio)
    date_str = selected_date.strftime("%Y%m%d")
    n_rec = int(len(df_recursos) if df_recursos is not None else 0)
    fname = f"lanc_recursos_{conv_slug}_{date_str}_n{n_rec}_total{total_recursos_pago:.2f}.csv"
    return fname, csv_bytes

def gerar_csv_nfse_lancamentos_bytes(
    df_nfse: pd.DataFrame,
    selected_convenio: str,
    referencia_str: Optional[str] = None,
    encoding: str = "cp1252",
):
    """
    Gera CSV de lançamentos contábeis para o Relatório NFS-e.

    Regra pedida:
      - Para cada linha/remessa do relatório:
          • 1 lançamento na conta 1133001 (Débito / +)
          • 1 lançamento na conta 1133003 (Crédito / -)
        Usando a MESMA lógica de subconta dos demais CSVs (SUBCONTA_MAP + overrides).
        Descrição:
          "Rem. (Nº Remessa) - NFSe (Nº NF) - Fat. (Ref.)"
    """
    import io, csv as _csv, unicodedata

    # ---- prefs / overrides / unidade (igual outras funções de CSV) ----
    prefs: dict = {}
    ovrs: dict = {}
    unidade: str = ""
    if st is not None:
        try:
            prefs = st.session_state.get("csv_prefs", {}) or {}
            ovrs = st.session_state.get("csv_convenio_overrides", {}) or {}
            unidade = st.session_state.get("unidade", "") or ""
        except Exception:
            prefs, ovrs, unidade = {}, {}, ""

    oconv: dict = {}
    if isinstance(ovrs, dict):
        oconv = ovrs.get(selected_convenio, {}) or {}

    header_first_row = prefs.get("header_first_row")
    if not header_first_row:
        header_first_row = (
            ("1941", "Clínica Adventista de Curitiba - IASBS")
            if (unidade or "").upper() == "CMAC"
            else ("1841", "Clínica Adventista de Porto Alegre - IASBS")
        )

    # Mesmo SUBCONTA_MAP usado nas outras funções
    SUBCONTA_MAP = {
        "Afpergs": "13",
        "Proasa": "51",
        "Amil": "39",
        "Assefaz": "27",
        "Banco Central": "43",
        "Cassi": "26",
        "Doctor": "54",
        "Embratel": "18",
        "Humana": "55",
        "GEAP": "19",
        "Geap": "19",
        "Postal Saúde": "29",
        "Prevent Senior": "91",
        "Saúde Caixa": "15",
        "Medservice": "16",
        "Life": "45",
        "Gente Saúde": "59",
        "Capesesp": "20",
        "Ipê Saúde": "12",
    }

    subconta_convenio = (
        oconv.get("subconta_convenio")
        or SUBCONTA_MAP.get(selected_convenio, "13")
    )

    def _nfc(s):
        return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

    def _to_float(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        try:
            return float(v)
        except Exception:
            s = str(v).strip().replace("R$", "").replace(" ", "")
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            try:
                return float(s)
            except Exception:
                return None

    # ---- se não tiver dados, só devolve o cabeçalho ----
    rows = [_ensure_len(list(header_first_row), CSV_COLS)]
    if df_nfse is None or len(df_nfse) == 0:
        sio = io.StringIO()
        w = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
        w.writerows(rows)
        csv_bytes = sio.getvalue().encode(encoding, errors="replace")
        conv_slug = _slugify(selected_convenio)
        fname = f"lanc_nfse_{conv_slug}_vazio.csv"
        return fname, csv_bytes

    total_debito = 0.0
    n_itens = 0

    # Garante colunas esperadas
    for c in ["Nº Remessa", "Nº NF", "NF recurso", "Ref.", "Valor NF"]:
        if c not in df_nfse.columns:
            df_nfse[c] = None

    for _, r in df_nfse.iterrows():
        num_remessa = str(r.get("Nº Remessa") or "").strip()

        # Número de NF a usar na descrição:
        # - Para "RG" (NF encontrada em NF recurso), usa o campo "NF recurso"
        # - Caso contrário, usa "Nº NF" (padrão)
        nf_normal = str(r.get("Nº NF") or "").strip()
        nf_recurso = str(r.get("NF recurso") or "").strip()
        match_kind = str(r.get("_nfse_match_kind") or "").strip().lower()
        ref_raw = str(r.get("Ref.") or "").strip()
        ref_fmt = _fmt_ref_mmYYYY(ref_raw)

        is_rg = (match_kind in ("recurso", "rg")) or ref_raw.strip().lower().startswith("rg")
        nf_to_use = nf_recurso if (is_rg and nf_recurso) else (nf_normal or nf_recurso)

        # Valor a lançar:
        # - RG => Valor recursado
        # - normal => Valor NF
        if is_rg:
            valor_lanc = _as_number(r.get("Valor recursado"))
        else:
            valor_lanc = _as_number(r.get("Valor NF"))

        valor_lanc = float(valor_lanc or 0.0)
        if valor_lanc == 0:
            continue

        # Descrição:
        # Normal: "Rem. (N° Remessa) - NFSe (NFSe) - Fat. (Ref.)"
        # RG:     "... - NFSe (NF recurso) - Rg. (Ref.)"
        if is_rg:
            desc = f"Rem. {num_remessa} - NFSe {nf_to_use} - Rg. {ref_fmt}"
        else:
            desc = f"Rem. {num_remessa} - NFSe {nf_to_use} - Fat. {ref_fmt}"

        # Contas:
        # - normal: débito 1133001 / crédito 1133003
        # - RG:     débito 1133001 / crédito 1133004
        conta_credito = "1133004" if is_rg else "1133003"

        rows.append(_ensure_len([
            "1133001", subconta_convenio, "10", "1110", "0A",
            _fmt_amount_csv(valor_lanc), "N", _nfc(desc)
        ], CSV_COLS))

        rows.append(_ensure_len([
            conta_credito, subconta_convenio, "10", "1110", "0A",
            "-" + _fmt_amount_csv(valor_lanc), "N", _nfc(desc)
        ], CSV_COLS))

    # ---- grava CSV ----
    sio = io.StringIO()
    writer = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
    writer.writerows(rows)
    csv_text = sio.getvalue()
    csv_bytes = csv_text.encode(encoding, errors="replace")

    conv_slug = _slugify(selected_convenio)
    ref_slug = _slugify(referencia_str) if referencia_str else ""
    total_str = f"{float(total_debito):.2f}"

    if ref_slug:
        fname = f"lanc_nfse_{conv_slug}_{ref_slug}_n{n_itens}_total{total_str}.csv"
    else:
        fname = f"lanc_nfse_{conv_slug}_n{n_itens}_total{total_str}.csv"

    return fname, csv_bytes

def gerar_csv_lancamentos_unimed_bytes(
    payload: dict,
    encoding: str = "cp1252",
):
    """
    Regras:
    - Débito: 1136001 / sub 1141, um lançamento por entidade com o SOMATÓRIO
      do 'Valor Reembolsado' dos itens daquela entidade.
      Descrição: "{Entidade} - Depósito C/c - Unimed"

    - Crédito: 1131001 / sub 1122, um lançamento por item (nome) daquela entidade,
      com o valor do 'Valor Reembolsado' (negativo).
      Descrição: 'NFSe {Número Nota Fiscal} - {Titular}'
    """
    import io, csv as _csv, unicodedata

    # Pega header_first_row do modelo padrão do teu sistema (igual as outras rotinas)
    prefs: dict = {}
    unidade: str = ""
    if st is not None:
        try:
            prefs = st.session_state.get("csv_prefs", {}) or {}
            unidade = st.session_state.get("unidade", "") or ""
        except Exception:
            prefs, unidade = {}, ""

    unid = (unidade or "").strip().upper()

    deb_subconta = "1141"

    if unid == "CMAP":
        cred_subconta = "1121"
    else:
        # padrão CMAC (e qualquer outro caso)
        cred_subconta = "1122"

    header_first_row = prefs.get("header_first_row")
    if not header_first_row:
        # mesmo padrão usado no resto do backend.py
        if (unidade or "").upper() == "CMAC":
            header_first_row = ("1941", "Clínica Adventista de Curitiba - IASBS")
        else:
            header_first_row = ("1841", "Clínica Adventista de Porto Alegre - IASBS")

    def _nfc(s):
        return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

    def _to_float(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        try:
            return float(v)
        except Exception:
            s = str(v).strip().replace("R$", "").replace(" ", "")
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            try:
                return float(s)
            except Exception:
                return 0.0

    def _formatar_nome_titulo(nome: str) -> str:
        """
        Primeira letra maiúscula em cada palavra,
        mantendo conectores em minúsculo.
        """
        if not nome:
            return ""

        conectores = {
            "de", "da", "do", "das", "dos", "e"
        }

        partes = nome.strip().lower().split()
        resultado = []

        for p in partes:
            if p in conectores:
                resultado.append(p)
            else:
                resultado.append(p.capitalize())

        return " ".join(resultado)


    items = payload.get("items") or []
    if not items:
        # devolve só header, arquivo vazio
        rows = [_ensure_len(list(header_first_row), CSV_COLS)]
        sio = io.StringIO()
        w = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
        w.writerows(rows)
        return "lanc_unimed_vazio.csv", sio.getvalue().encode(encoding, errors="replace")

    # Monta DataFrame a partir do JSON (pega campos que você citou)
    base_rows = []
    for it in items:
        entidade = str(it.get("Entidade") or "").strip()
        titular = str(it.get("Titular") or "").strip()

        row_xlsx = it.get("row_xlsx") or {}
        nf = row_xlsx.get("Número Nota Fiscal")
        valor = row_xlsx.get("Valor Reembolsado")

        base_rows.append({
            "Entidade": entidade,
            "Titular": titular,
            "NumeroNF": "" if nf is None else str(nf).strip(),
            "ValorReembolsado": _to_float(valor),
        })

    df = pd.DataFrame(base_rows)
    df = df[df["Entidade"].astype(str).str.strip() != ""].copy()
    df = df[df["ValorReembolsado"] != 0].copy()

    # Header do arquivo
    rows = [_ensure_len(list(header_first_row), CSV_COLS)]

    # 1) DÉBITO por entidade (somatório)
    debitos = df.groupby("Entidade", dropna=False)["ValorReembolsado"].sum().reset_index()
    for _, r in debitos.iterrows():
        ent = str(r["Entidade"]).strip()
        total = float(r["ValorReembolsado"] or 0.0)
        if total == 0:
            continue

        desc = f"{ent} - Depósito C/c - Unimed"
        rows.append(_ensure_len([
            "1136001", deb_subconta, "10", "1110", "0A",
            _fmt_amount_csv(total), "N", _nfc(desc)
        ], CSV_COLS))

        # 2) CRÉDITO item a item (por pessoa) para essa entidade
        df_ent = df[df["Entidade"] == ent]
        for _, ri in df_ent.iterrows():
            valor_i = float(ri["ValorReembolsado"] or 0.0)
            if valor_i == 0:
                continue
            nf_i = str(ri["NumeroNF"] or "").strip()
            tit_raw = str(ri["Titular"] or "").strip()
            tit_fmt = _formatar_nome_titulo(tit_raw)
            desc_i = f"NFSe {nf_i} - {tit_fmt}".strip()

            rows.append(_ensure_len([
                "1131001", cred_subconta, "10", "1110", "0A",
                "-" + _fmt_amount_csv(valor_i), "N", _nfc(desc_i)
            ], CSV_COLS))

    # Gera bytes
    sio = io.StringIO()
    w = _csv.writer(sio, delimiter=CSV_DELIM, lineterminator="\r\n")
    w.writerows(rows)
    csv_bytes = sio.getvalue().encode(encoding, errors="replace")

    # Nome do arquivo
    total_geral = float(df["ValorReembolsado"].sum() or 0.0)
    fname = f"lanc_unimed_total{total_geral:.2f}.csv"
    return fname, csv_bytes

# --------------------------------------------------------------------------- #
# NOVO: Identificação Unimed (XLSX + CSV -> JSON)
# --------------------------------------------------------------------------- #

def processar_identificacao_unimed_para_json(
    xlsx_file,
    csv_file,
    output_dir: str = ".",
    threshold: float = 0.78,
) -> str:
    """
    Lê:
      - XLSX: cabeçalho em A5:F5 (header=4 no pandas), colunas incluem "Titular" e "Entidade"
      - CSV: colunas incluem "Nome" e "entidade"/"Entidade"

    Regra:
      - Para cada Titular (XLSX), encontra o Nome (CSV) mais similar.
      - Se score >= threshold, escreve o número da "entidade" do CSV em "Entidade" do XLSX.
      - Gera JSON estruturado com matches.
    """
    import os, json, re
    from difflib import SequenceMatcher
    import pandas as pd
    import numpy as np
    import datetime as dt

    os.makedirs(output_dir, exist_ok=True)

    # ---------- helpers ----------
    def _norm_name(s: object) -> str:
        s = _norm(s)
        s = re.sub(r"[^a-z0-9 ]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        # token sort p/ melhorar matching (ex: "JOAO DA SILVA" vs "SILVA JOAO")
        toks = [t for t in s.split(" ") if t]
        toks.sort()
        return " ".join(toks)

    def _ratio(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        return SequenceMatcher(None, a, b).ratio()

    def _find_col(df: pd.DataFrame, candidates: list[str], fallback_idx: int | None = None) -> str:
        cols = list(df.columns)
        norm_cols = [_norm(c) for c in cols]
        for cand in candidates:
            nc = _norm(cand)
            for i, c in enumerate(norm_cols):
                if c == nc:
                    return cols[i]
        if fallback_idx is not None and 0 <= fallback_idx < len(cols):
            return cols[fallback_idx]
        return cols[0] if cols else ""
    
    def _to_json_safe(v):
        # pandas / numpy nulos
        try:
            import pandas as pd
            if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                return None
        except Exception:
            pass

        # Timestamp (pandas) e datetime/date -> string ISO
        try:
            import pandas as pd
            if isinstance(v, pd.Timestamp):
                # NaT
                if pd.isna(v):
                    return None
                return v.isoformat()
        except Exception:
            pass

        import datetime as dt
        if isinstance(v, (dt.datetime, dt.date)):
            return v.isoformat()

        # numpy scalars -> python scalars
        try:
            import numpy as np
            if isinstance(v, np.generic):
                return v.item()
        except Exception:
            pass

        # fallback
        return v


    # ---------- XLSX ----------
    # cabeçalho na linha 5 => header=4 (0-based)
    df_xlsx = pd.read_excel(xlsx_file, header=4, engine="openpyxl")
    df_xlsx = df_xlsx.dropna(how="all")

    col_titular = _find_col(df_xlsx, ["Titular"], fallback_idx=0)
    col_entidade_xlsx = _find_col(df_xlsx, ["Entidade"], fallback_idx=1)

    # ---------- CSV (tenta auto-sep; cai p/ ';') ----------
    try:
        df_csv = pd.read_csv(csv_file, sep=None, engine="python")
    except Exception:
        df_csv = pd.read_csv(csv_file, sep=";", engine="python")

    df_csv = df_csv.dropna(how="all")
    df_csv.columns = [str(c).replace("\ufeff", "").strip() for c in df_csv.columns]

    col_nome = _find_col(df_csv, ["Nome"], fallback_idx=0)
    col_entidade_csv = _find_col(df_csv, ["entidade", "Entidade"], fallback_idx=1)

    # prepara lista de nomes do CSV
    csv_rows = []
    for _, r in df_csv.iterrows():
        nome_raw = r.get(col_nome)
        ent = r.get(col_entidade_csv)
        nome_n = _norm_name(nome_raw)
        if nome_n:
            csv_rows.append((nome_n, str(ent).strip() if ent is not None else ""))

    # ---------- matching ----------
    out_items = []
    entidades_preenchidas = 0

    # garante coluna "Entidade" no df_xlsx
    if col_entidade_xlsx not in df_xlsx.columns:
        df_xlsx[col_entidade_xlsx] = ""

    for idx, row in df_xlsx.iterrows():
        titular_raw = row.get(col_titular)
        titular_n = _norm_name(titular_raw)

        best_score = 0.0
        best_ent = ""
        best_nome_norm = ""

        if titular_n and csv_rows:
            for nome_norm, ent in csv_rows:
                sc = _ratio(titular_n, nome_norm)
                if sc > best_score:
                    best_score = sc
                    best_ent = ent
                    best_nome_norm = nome_norm

        matched = (best_score >= float(threshold)) and (best_ent != "")
        if matched:
            df_xlsx.at[idx, col_entidade_xlsx] = best_ent
            entidades_preenchidas += 1

        out_items.append({
            "linha_xlsx": int(idx) + 6,  # aproximação visual (dados começam após header)
            "Titular": None if pd.isna(titular_raw) else str(titular_raw),
            "Entidade": str(best_ent) if matched else ("" if pd.isna(row.get(col_entidade_xlsx)) else str(row.get(col_entidade_xlsx))),
            "match": {
                "encontrou": bool(matched),
                "score": round(float(best_score), 4),
                "nome_csv_normalizado": best_nome_norm,
                "threshold": float(threshold),
            },
            "row_xlsx": {k: _to_json_safe(v) for k, v in row.to_dict().items()},
        })

    payload = {
        "meta": {
            "threshold": float(threshold),
            "total_linhas_xlsx": int(len(df_xlsx)),
            "total_linhas_csv": int(len(df_csv)),
            "entidades_preenchidas": int(entidades_preenchidas),
            "colunas": {
                "xlsx_titular": col_titular,
                "xlsx_entidade": col_entidade_xlsx,
                "csv_nome": col_nome,
                "csv_entidade": col_entidade_csv,
            },
        },
        "items": out_items,
    }

    fpath = os.path.join(output_dir, "saida_identificacao_unimed.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return fpath

# CABERGS #
from io import BytesIO
import pandas as pd

def _excel_col_idx(col_letters: str) -> int:
    """A=1, B=2 ..."""
    col_letters = col_letters.upper().strip()
    n = 0
    for ch in col_letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def _safe_strip_lower(x) -> str:
    return str(x or "").strip().lower()

def processar_cabergs_arquivos(uploaded_files) -> pd.DataFrame:
    """
    Recebe lista de UploadedFile do Streamlit.
    - Corrige cabeçalho na linha 11:
      B -> C, AB -> AF, AQ -> AR
    - Extrai:
      G5 (remessa) e X5 (competência)
    - Retorna um DataFrame consolidado para exibir na UI.
    """
    if not uploaded_files:
        return pd.DataFrame()

    frames = []

    for uf in uploaded_files:
        name = getattr(uf, "name", "arquivo")
        raw = uf.read()
        ext = name.lower().split(".")[-1]

        # -----------------------
        # Caso XLSX: openpyxl
        # -----------------------
        if ext == "xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
                ws = wb.active

                # valores importantes
                remessa = ws["G5"].value
                competencia = ws["X5"].value

                # corrigir cabeçalho (linha 11)
                row_header = 11
                # B -> C
                ws.cell(row_header, _excel_col_idx("C")).value = ws.cell(row_header, _excel_col_idx("B")).value
                ws.cell(row_header, _excel_col_idx("B")).value = None
                # AB -> AF
                ws.cell(row_header, _excel_col_idx("AF")).value = ws.cell(row_header, _excel_col_idx("AB")).value
                ws.cell(row_header, _excel_col_idx("AB")).value = None
                # AQ -> AR
                ws.cell(row_header, _excel_col_idx("AR")).value = ws.cell(row_header, _excel_col_idx("AQ")).value
                ws.cell(row_header, _excel_col_idx("AQ")).value = None

                # ler headers a partir da coluna B (como você disse)
                start_col = _excel_col_idx("B")
                max_col = ws.max_column

                headers = []
                used = {}

                for c in range(start_col, max_col + 1):
                    raw = ws.cell(row_header, c).value

                    # trata None, vazio e 'nan' como sem nome
                    if raw is None:
                        name = f"COL_{c}"
                    else:
                        s = str(raw).strip()
                        if s == "" or s.lower() == "nan":
                            name = f"COL_{c}"
                        else:
                            name = s

                    headers.append(name)

                # garante unicidade de cabeçalhos (evita 'nan' repetido, etc.)
                headers = _dedupe_headers(headers)


                # colunas que vamos ler (B..max_col)
                col_indices = list(range(start_col, max_col + 1))

                # montar linhas (a partir da 12)
                data_rows = []
                for r in range(row_header + 1, ws.max_row + 1):
                    row_vals = [ws.cell(r, c).value for c in col_indices]
                    # pula linha vazia
                    if all(v is None or str(v).strip() == "" for v in row_vals):
                        continue
                    data_rows.append(row_vals)

                df = pd.DataFrame(data_rows, columns=headers)

                df.columns = _dedupe_headers(df.columns.tolist())

                # limpa colunas vazias e linhas sem 'Início'
                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{
                    "Arquivo": name,
                    "Erro": f"Falha ao ler XLSX: {e}"
                }])
                remessa = None
                competencia = None

        # -----------------------
        # Caso XLS: pandas (xlrd)
        # -----------------------
        elif ext == "xls":
            try:
                # depende de xlrd instalado. se não estiver, vai cair no except.
                raw_df = pd.read_excel(BytesIO(raw), header=None, engine="xlrd")

                # G5 = linha 5 col G
                remessa = raw_df.iat[4, _excel_col_idx("G") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("G") - 1) else None
                # X5 = linha 5 col X
                competencia = raw_df.iat[4, _excel_col_idx("X") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("X") - 1) else None

                # corrigir cabeçalho na linha 11 (índice 10)
                hr = 10
                # B -> C
                raw_df.iat[hr, _excel_col_idx("C") - 1] = raw_df.iat[hr, _excel_col_idx("B") - 1]
                raw_df.iat[hr, _excel_col_idx("B") - 1] = None
                # AB -> AF
                raw_df.iat[hr, _excel_col_idx("AF") - 1] = raw_df.iat[hr, _excel_col_idx("AB") - 1]
                raw_df.iat[hr, _excel_col_idx("AB") - 1] = None
                # AQ -> AR
                raw_df.iat[hr, _excel_col_idx("AR") - 1] = raw_df.iat[hr, _excel_col_idx("AQ") - 1]
                raw_df.iat[hr, _excel_col_idx("AQ") - 1] = None

                # cabeçalho começa na coluna B
                start_c = _excel_col_idx("B") - 1
                headers = raw_df.iloc[hr, start_c:].tolist()
                headers = [
                    f"COL_{start_c+i+1}" if (h is None or str(h).strip() == "") else str(h).strip()
                    for i, h in enumerate(headers)
                ]

                df = raw_df.iloc[hr+1:, start_c:].copy()
                headers = _dedupe_headers(headers)
                df.columns = headers
                df = df.dropna(how="all")

                # limpa colunas vazias e linhas sem 'Início'
                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{
                    "Arquivo": name,
                    "Erro": f"Falha ao ler XLS: {e} (se der isso, eu ajusto pra você)"
                }])
                remessa = None
                competencia = None

        else:
            df = pd.DataFrame([{"Arquivo": name, "Erro": "Extensão não suportada"}])
            remessa = None
            competencia = None

        # metadados do arquivo (evita colunas duplicadas)
        for meta_col in ["Arquivo", "Remessa (G5)", "Competência (X5)"]:
            if meta_col in df.columns:
                df = df.rename(columns={meta_col: f"{meta_col} (orig)"})

        df.insert(0, "Arquivo", name)
        df.insert(1, "Remessa (G5)", remessa)
        df.insert(2, "Competência (X5)", competencia)

        # garantia final: nomes únicos
        df.columns = _dedupe_headers(df.columns.tolist())

        # limpa colunas vazias e linhas sem 'Início'
        df = _drop_nan_only_columns(df)
        df = _drop_rows_without_inicio(df, "Início")

        frames.append(df)

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # limpeza final (caso algum arquivo traga colunas/linhas vazias)
    out = _drop_nan_only_columns(out)
    out = _drop_rows_without_inicio(out, "Início")

    return out


# =============================================================================
# CSV (análise) + Conciliação CABERGS x CSV
# =============================================================================

import io
import csv
import re

def _read_csv_robusto(uploaded_file) -> pd.DataFrame:
    """Lê CSV com fallback de encoding e delimitador, sem depender de sep=None."""
    if uploaded_file is None:
        return pd.DataFrame()

    raw = uploaded_file.read()
    if not raw:
        return pd.DataFrame()

    # tenta decodificar
    text = None
    last_err = None
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(enc)
            break
        except Exception as e:
            last_err = e
    if text is None:
        raise RuntimeError(f"Falha ao decodificar CSV: {last_err}")

    sample = text[:20000]
    # tenta sniffer
    sniffer_delim = None
    try:
        sniffer_delim = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"]).delimiter
    except Exception:
        sniffer_delim = None

    counts = {
        ";": sample.count(";"),
        ",": sample.count(","),
        "\t": sample.count("\t"),
        "|": sample.count("|"),
    }
    best_by_count = max(counts, key=counts.get)

    seps_to_try = []
    if sniffer_delim:
        seps_to_try.append(sniffer_delim)
    seps_to_try.append(best_by_count)
    for s in [";", ",", "\t", "|"]:
        if s not in seps_to_try:
            seps_to_try.append(s)

    df = None
    last_read_err = None
    for sep in seps_to_try:
        try:
            df_try = pd.read_csv(io.StringIO(text), sep=sep, engine="python")
            # se ficou 1 coluna só e o sep aparece bastante, tenta próximo
            if df_try.shape[1] == 1 and counts.get(sep, 0) > 0:
                df = df_try  # guarda, mas tenta outra
                continue
            df = df_try
            break
        except Exception as e:
            last_read_err = e
            df = None

    if df is None:
        raise RuntimeError(f"Falha ao ler CSV. Último erro: {last_read_err}")

    # remove colunas lixo comuns
    df = df.loc[:, [
        c for c in df.columns
        if not str(c).lower().startswith("unnamed")
        and not str(c).startswith("\ufeff")
    ]]

    return df

def processar_csv_analise(uploaded_file) -> pd.DataFrame:
    """Tratamento inicial do CSV (limpeza + normalizações leves)."""
    df = _read_csv_robusto(uploaded_file)

    if df is None or df.empty:
        return pd.DataFrame()

    # remove colunas só com nan/vazio
    df = _drop_nan_only_columns(df)

    # remove linhas totalmente vazias
    df = df.dropna(how="all")

    # limpa colunas vazias e linhas sem 'Início'
    df = _drop_nan_only_columns(df)
    df = _drop_rows_without_inicio(df, "Início")

    cols_to_drop = [
    "Seq Proced",
    "Cod Tab",
    "Qtd Cob",
    "Unit Cob",
    "Qtd Glosa",
    "Unit Glosa",
    "Qtd Pg",
    "Uso",
    "Descrição Uso",
    ]

    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors="ignore")

    # -------------------------------------------------------
    # Cria coluna "Valor MV" ao lado esquerdo de "Total Pago"
    # -------------------------------------------------------
    if "Valor MV" not in df.columns:
        df["Valor MV"] = None  # ou 0.0, se preferir numérico

    if "Total Pago" in df.columns:
        cols = df.columns.tolist()
        cols.remove("Valor MV")
        idx = cols.index("Total Pago")
        cols.insert(idx, "Valor MV")  # entra à esquerda do Total Pago
        df = df[cols]

    # trim strings
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip().replace({"nan": None, "None": None, "": None})

    return df.reset_index(drop=True)

def _only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def _norm_guia_tiss_tabela1(v) -> str:
    """Tabela 1 (XLS): se começar com 0, desconsidera zeros à esquerda."""
    s = "" if v is None else str(v).strip()
    s = _only_digits(s)
    # remove zeros à esquerda (regra)
    s = s.lstrip("0")
    return s

def _norm_guia_tiss_tabela2(v) -> str:
    """Tabela 2 (CSV): remove 3 primeiros caracteres e o último, depois mantém só dígitos."""
    s = "" if v is None else str(v).strip()
    if len(s) >= 5:
        s = s[3:-1]
    else:
        s = ""
    s = _only_digits(s)
    return s

def _find_col_case_insensitive(df: pd.DataFrame, target: str) -> str | None:
    if df is None or df.empty:
        return None
    tgt = (target or "").strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == tgt:
            return c
    return None



def gerar_tabela_conciliacao(df_xls: pd.DataFrame, df_csv: pd.DataFrame) -> pd.DataFrame:
    """
    Gera a 3ª tabela (conciliação) cruzando Guia TISS (tabela1) com Guia Tiss (tabela2)
    após normalizações definidas pelo usuário.
    """
    if df_xls is None:
        df_xls = pd.DataFrame()
    if df_csv is None:
        df_csv = pd.DataFrame()

    if df_xls.empty:
        return pd.DataFrame()
    if df_csv.empty:
        # ainda assim retorna com status não encontrado
        df_base = df_xls.copy()
        c1 = _find_col_case_insensitive(df_base, "Guia TISS") or "Guia TISS"
        if c1 not in df_base.columns:
            df_base[c1] = None
        df_base["_key_guia"] = df_base[c1].apply(_norm_guia_tiss_tabela1)
        df_base["Status"] = "Não encontrado"
        return df_base

    col_xls = _find_col_case_insensitive(df_xls, "Guia TISS") or "Guia TISS"
    col_csv = _find_col_case_insensitive(df_csv, "Guia Tiss") or _find_col_case_insensitive(df_csv, "Guia TISS") or "Guia Tiss"

    dfx = df_xls.copy()
    if col_xls not in dfx.columns:
        dfx[col_xls] = None
    dfx["_key_guia"] = dfx[col_xls].apply(_norm_guia_tiss_tabela1)

    dfc = df_csv.copy()
    if col_csv not in dfc.columns:
        dfc[col_csv] = None
    dfc["_key_guia"] = dfc[col_csv].apply(_norm_guia_tiss_tabela2)

    # remove chaves vazias
    dfx = dfx[dfx["_key_guia"].astype(str).str.strip() != ""].copy()
    dfc = dfc[dfc["_key_guia"].astype(str).str.strip() != ""].copy()

    # se houver múltiplas linhas no CSV por chave, agregamos somando colunas numéricas
    num_cols = [c for c in dfc.columns if c != "_key_guia" and pd.api.types.is_numeric_dtype(dfc[c])]
    if num_cols:
        dfc_agg = dfc.groupby("_key_guia", as_index=False)[num_cols].sum()
        # pega a primeira ocorrência dos campos texto
        text_cols = [c for c in dfc.columns if c not in num_cols and c != "_key_guia"]
        if text_cols:
            dfc_first = dfc.groupby("_key_guia", as_index=False)[text_cols].first()
            dfc = pd.merge(dfc_first, dfc_agg, on="_key_guia", how="left")
        else:
            dfc = dfc_agg
    else:
        # dedupe simples
        dfc = dfc.drop_duplicates(subset=["_key_guia"]).copy()

    merged = dfx.merge(
        dfc,
        on="_key_guia",
        how="left",
        suffixes=("_XLS", "_CSV"),
        indicator=True
    )

    merged["Status"] = merged["_merge"].map({"both": "Encontrado", "left_only": "Não encontrado", "right_only": "(extra CSV)"})
    merged = merged.drop(columns=["_merge"])

    # Reorganiza colunas para ficar bom de analisar
    cols_front = []
    # status + chaves
    cols_front += ["Status", "_key_guia"]
    # guia original
    if col_xls in merged.columns:
        cols_front.append(col_xls)
    if col_csv in merged.columns:
        cols_front.append(col_csv)

    # tenta trazer campos úteis comuns da tabela1
    for c in ["Atendimento", "Nr.Conta", "Início", "Término", "Origem", "Valor"]:
        c_real = _find_col_case_insensitive(merged, c)
        if c_real and c_real not in cols_front:
            cols_front.append(c_real)

    # resto das colunas
    rest = [c for c in merged.columns if c not in cols_front]
    merged = merged[cols_front + rest]

    return merged.reset_index(drop=True)




def marcar_encontrados_csv(
    df_xls: pd.DataFrame,
    df_csv: pd.DataFrame,
    flag_col: str = "Encontrado",
    remessa_col_out: str = "Remessa (G5)",
    col_xls_nome: str = "Guia TISS",
    col_csv_nome: str = "Guia Tiss",
    col_csv_data_atend: str = "Data Atend",
    col_xls_termino: str = "Término",
) -> pd.DataFrame:
    """
    Marca no CSV (Tabela 2) os registros encontrados no XLS (Tabela 1) via Guia TISS normalizado.
    Também preenche uma coluna de remessa no CSV.

    Regra adicional (remessa exata):
    - Se a data do atendimento no CSV (col_csv_data_atend) bater com a data de término no XLS (col_xls_termino)
      para aquela mesma chave (Guia), então escolhe a remessa correspondente a essa data (um único número).
    - Se não bater, mantém o comportamento anterior: concatena remessas únicas (ex: "20735, 20736").
    """
    if df_csv is None:
        return pd.DataFrame()

    out = df_csv.copy()

    # garante coluna checkbox no começo
    if flag_col not in out.columns:
        out.insert(0, flag_col, False)
    else:
        out[flag_col] = out[flag_col].fillna(False).astype(bool)

    # garante coluna de remessa logo após a checkbox
    if remessa_col_out not in out.columns:
        out.insert(1, remessa_col_out, "")
    else:
        out[remessa_col_out] = out[remessa_col_out].fillna("").astype(str)

    if df_xls is None or df_xls.empty or out.empty:
        return out

    col_xls = _find_col_case_insensitive(df_xls, col_xls_nome) or col_xls_nome
    col_csv = (
        _find_col_case_insensitive(out, col_csv_nome)
        or _find_col_case_insensitive(out, "Guia TISS")
        or col_csv_nome
    )

    col_remessa_xls = _find_col_case_insensitive(df_xls, "Remessa (G5)") or "Remessa (G5)"
    col_termino_xls = _find_col_case_insensitive(df_xls, col_xls_termino) or col_xls_termino
    col_data_csv = _find_col_case_insensitive(out, col_csv_data_atend) or col_csv_data_atend

    if col_xls not in df_xls.columns or col_csv not in out.columns:
        return out

    # -------------------------
    # prepara tabela XLS com chave + (termino, remessa)
    # -------------------------
    dfx = df_xls.copy()

    if col_remessa_xls not in dfx.columns:
        dfx[col_remessa_xls] = ""
    if col_termino_xls not in dfx.columns:
        dfx[col_termino_xls] = None

    dfx["_key_guia"] = dfx[col_xls].apply(_norm_guia_tiss_tabela1).astype(str).str.strip()
    dfx = dfx[dfx["_key_guia"] != ""]

    # parse término
    dfx["_termino_dt"] = pd.to_datetime(dfx[col_termino_xls], errors="coerce", dayfirst=True).dt.date

    # normaliza remessa
    def _norm_rem(x):
        s = "" if x is None else str(x).strip()
        if s.lower() in ("nan", "none"):
            return ""
        return s

    dfx["_remessa_str"] = dfx[col_remessa_xls].apply(_norm_rem)
    dfx = dfx[dfx["_remessa_str"] != ""]

    # map chave -> lista de pares (termino_dt, remessa)
    pairs_map = (
        dfx.groupby("_key_guia")[["_termino_dt", "_remessa_str"]]
        .apply(lambda g: [(r["_termino_dt"], r["_remessa_str"]) for _, r in g.iterrows()])
        .to_dict()
    )

    # fallback: remessas únicas concatenadas
    rem_concat_map = {}
    for k, pairs in pairs_map.items():
        rems = sorted({p[1] for p in pairs if p[1]})
        rem_concat_map[k] = ", ".join(rems)

    keys_set = set(rem_concat_map.keys())

    # -------------------------
    # Regra NOVA: só marca/enche remessa se TISS + Data baterem
    # -------------------------
    keys_csv = out[col_csv].apply(_norm_guia_tiss_tabela2).astype(str).str.strip()

    # Data Atend (CSV) pode vir "07/10/2025 00:00:00" -> vira date(2025,10,7)
    if col_data_csv in out.columns:
        data_csv_dt = pd.to_datetime(out[col_data_csv], errors="coerce", dayfirst=True).dt.date
    else:
        data_csv_dt = pd.Series([None] * len(out))

    found_flags = []
    remessas_out = []

    for key, dt in zip(keys_csv.tolist(), data_csv_dt.tolist()):
        # precisa ter chave e data válida
        if not key or key not in pairs_map or dt is None:
            found_flags.append(False)
            remessas_out.append("")
            continue

        # procura remessa(s) no XLS cuja data Término == Data Atend
        matches = sorted({rem for (t, rem) in pairs_map[key] if t == dt and rem})

        if not matches:
            # TISS existe, mas data não bate -> não encontrado e sem remessa
            found_flags.append(False)
            remessas_out.append("")
            continue

        # achou pela data: encontrado e remessa exata (ou concat se tiver mais de uma)
        found_flags.append(True)
        remessas_out.append(matches[0] if len(matches) == 1 else ", ".join(matches))

    out[flag_col] = found_flags
    out[remessa_col_out] = remessas_out


    cols = [flag_col, remessa_col_out] + [c for c in out.columns if c not in (flag_col, remessa_col_out)]
    return out[cols]
