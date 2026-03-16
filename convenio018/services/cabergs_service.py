"""Fluxos auxiliares para CABERGS e conciliação CSV/XLS."""

from __future__ import annotations

import re
from io import BytesIO

import pandas as pd

from ..utils.normalizers import _dedupe_headers, _drop_nan_only_columns, _drop_rows_without_inicio, _find_col_case_insensitive, _norm_guia_tiss_tabela1, _norm_guia_tiss_tabela2, _only_digits, _safe_strip_lower
from ..utils.parsers import _excel_col_idx, _read_csv_robusto

def processar_cabergs_arquivos(uploaded_files) -> pd.DataFrame:
    """
    Recebe lista de UploadedFile do Streamlit.
    - Corrige cabeçalho na linha 11:
      B -> C, AB -> AF, AQ -> AR
    - Extrai:
      G5 (remessa) e X5 (competência)
    - Retorna um DataFrame consolidado para exibir na UI.
    """
    if not uploaded_files:
        return pd.DataFrame()

    frames = []

    for uf in uploaded_files:
        name = getattr(uf, "name", "arquivo")
        raw = uf.read()
        ext = name.lower().split(".")[-1]

        # -----------------------
        # Caso XLSX: openpyxl
        # -----------------------
        if ext == "xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
                ws = wb.active

                # valores importantes
                remessa = ws["G5"].value
                competencia = ws["X5"].value

                # corrigir cabeçalho (linha 11)
                row_header = 11
                # B -> C
                ws.cell(row_header, _excel_col_idx("C")).value = ws.cell(row_header, _excel_col_idx("B")).value
                ws.cell(row_header, _excel_col_idx("B")).value = None
                # AB -> AF
                ws.cell(row_header, _excel_col_idx("AF")).value = ws.cell(row_header, _excel_col_idx("AB")).value
                ws.cell(row_header, _excel_col_idx("AB")).value = None
                # AQ -> AR
                ws.cell(row_header, _excel_col_idx("AR")).value = ws.cell(row_header, _excel_col_idx("AQ")).value
                ws.cell(row_header, _excel_col_idx("AQ")).value = None

                # ler headers a partir da coluna B (como você disse)
                start_col = _excel_col_idx("B")
                max_col = ws.max_column

                headers = []
                used = {}

                for c in range(start_col, max_col + 1):
                    raw = ws.cell(row_header, c).value

                    # trata None, vazio e 'nan' como sem nome
                    if raw is None:
                        name = f"COL_{c}"
                    else:
                        s = str(raw).strip()
                        if s == "" or s.lower() == "nan":
                            name = f"COL_{c}"
                        else:
                            name = s

                    headers.append(name)

                # garante unicidade de cabeçalhos (evita 'nan' repetido, etc.)
                headers = _dedupe_headers(headers)


                # colunas que vamos ler (B..max_col)
                col_indices = list(range(start_col, max_col + 1))

                # montar linhas (a partir da 12)
                data_rows = []
                for r in range(row_header + 1, ws.max_row + 1):
                    row_vals = [ws.cell(r, c).value for c in col_indices]
                    # pula linha vazia
                    if all(v is None or str(v).strip() == "" for v in row_vals):
                        continue
                    data_rows.append(row_vals)

                df = pd.DataFrame(data_rows, columns=headers)

                df.columns = _dedupe_headers(df.columns.tolist())

                # limpa colunas vazias e linhas sem 'Início'
                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{
                    "Arquivo": name,
                    "Erro": f"Falha ao ler XLSX: {e}"
                }])
                remessa = None
                competencia = None

        # -----------------------
        # Caso XLS: pandas (xlrd)
        # -----------------------
        elif ext == "xls":
            try:
                # depende de xlrd instalado. se não estiver, vai cair no except.
                raw_df = pd.read_excel(BytesIO(raw), header=None, engine="xlrd")

                # G5 = linha 5 col G
                remessa = raw_df.iat[4, _excel_col_idx("G") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("G") - 1) else None
                # X5 = linha 5 col X
                competencia = raw_df.iat[4, _excel_col_idx("X") - 1] if raw_df.shape[0] > 4 and raw_df.shape[1] > (_excel_col_idx("X") - 1) else None

                # corrigir cabeçalho na linha 11 (índice 10)
                hr = 10
                # B -> C
                raw_df.iat[hr, _excel_col_idx("C") - 1] = raw_df.iat[hr, _excel_col_idx("B") - 1]
                raw_df.iat[hr, _excel_col_idx("B") - 1] = None
                # AB -> AF
                raw_df.iat[hr, _excel_col_idx("AF") - 1] = raw_df.iat[hr, _excel_col_idx("AB") - 1]
                raw_df.iat[hr, _excel_col_idx("AB") - 1] = None
                # AQ -> AR
                raw_df.iat[hr, _excel_col_idx("AR") - 1] = raw_df.iat[hr, _excel_col_idx("AQ") - 1]
                raw_df.iat[hr, _excel_col_idx("AQ") - 1] = None

                # cabeçalho começa na coluna B
                start_c = _excel_col_idx("B") - 1
                headers = raw_df.iloc[hr, start_c:].tolist()
                headers = [
                    f"COL_{start_c+i+1}" if (h is None or str(h).strip() == "") else str(h).strip()
                    for i, h in enumerate(headers)
                ]

                df = raw_df.iloc[hr+1:, start_c:].copy()
                headers = _dedupe_headers(headers)
                df.columns = headers
                df = df.dropna(how="all")

                # limpa colunas vazias e linhas sem 'Início'
                df = _drop_nan_only_columns(df)
                df = _drop_rows_without_inicio(df, "Início")

            except Exception as e:
                df = pd.DataFrame([{
                    "Arquivo": name,
                    "Erro": f"Falha ao ler XLS: {e} (se der isso, eu ajusto pra você)"
                }])
                remessa = None
                competencia = None

        else:
            df = pd.DataFrame([{"Arquivo": name, "Erro": "Extensão não suportada"}])
            remessa = None
            competencia = None

        # metadados do arquivo (evita colunas duplicadas)
        for meta_col in ["Arquivo", "Remessa (G5)", "Competência (X5)"]:
            if meta_col in df.columns:
                df = df.rename(columns={meta_col: f"{meta_col} (orig)"})

        df.insert(0, "Arquivo", name)
        df.insert(1, "Remessa (G5)", remessa)
        df.insert(2, "Competência (X5)", competencia)

        # garantia final: nomes únicos
        df.columns = _dedupe_headers(df.columns.tolist())

        # limpa colunas vazias e linhas sem 'Início'
        df = _drop_nan_only_columns(df)
        df = _drop_rows_without_inicio(df, "Início")

        frames.append(df)

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # limpeza final (caso algum arquivo traga colunas/linhas vazias)
    out = _drop_nan_only_columns(out)
    out = _drop_rows_without_inicio(out, "Início")

    return out


def processar_csv_analise(uploaded_file) -> pd.DataFrame:
    """Tratamento inicial do CSV (limpeza + normalizações leves)."""
    df = _read_csv_robusto(uploaded_file)

    if df is None or df.empty:
        return pd.DataFrame()

    # remove colunas só com nan/vazio
    df = _drop_nan_only_columns(df)

    # remove linhas totalmente vazias
    df = df.dropna(how="all")

    # limpa colunas vazias e linhas sem 'Início'
    df = _drop_nan_only_columns(df)
    df = _drop_rows_without_inicio(df, "Início")

    cols_to_drop = [
    "Seq Proced",
    "Cod Tab",
    "Qtd Cob",
    "Unit Cob",
    "Qtd Glosa",
    "Unit Glosa",
    "Qtd Pg",
    "Uso",
    "Descrição Uso",
    ]

    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors="ignore")

    # -------------------------------------------------------
    # Cria coluna "Valor MV" ao lado esquerdo de "Total Pago"
    # -------------------------------------------------------
    if "Valor MV" not in df.columns:
        df["Valor MV"] = None  # ou 0.0, se preferir numérico

    if "Total Pago" in df.columns:
        cols = df.columns.tolist()
        cols.remove("Valor MV")
        idx = cols.index("Total Pago")
        cols.insert(idx, "Valor MV")  # entra à esquerda do Total Pago
        df = df[cols]

    # trim strings
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip().replace({"nan": None, "None": None, "": None})

    return df.reset_index(drop=True)

def _only_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def _norm_guia_tiss_tabela1(v) -> str:
    """Tabela 1 (XLS): se começar com 0, desconsidera zeros à esquerda."""
    s = "" if v is None else str(v).strip()
    s = _only_digits(s)
    # remove zeros à esquerda (regra)
    s = s.lstrip("0")
    return s

def _norm_guia_tiss_tabela2(v) -> str:
    """Tabela 2 (CSV): remove 3 primeiros caracteres e o último, depois mantém só dígitos."""
    s = "" if v is None else str(v).strip()
    if len(s) >= 5:
        s = s[3:-1]
    else:
        s = ""
    s = _only_digits(s)
    return s

def _find_col_case_insensitive(df: pd.DataFrame, target: str) -> str | None:
    if df is None or df.empty:
        return None
    tgt = (target or "").strip().lower()
    for c in df.columns:
        if str(c).strip().lower() == tgt:
            return c
    return None



def gerar_tabela_conciliacao(df_xls: pd.DataFrame, df_csv: pd.DataFrame) -> pd.DataFrame:
    """
    Gera a 3ª tabela (conciliação) cruzando Guia TISS (tabela1) com Guia Tiss (tabela2)
    após normalizações definidas pelo usuário.
    """
    if df_xls is None:
        df_xls = pd.DataFrame()
    if df_csv is None:
        df_csv = pd.DataFrame()

    if df_xls.empty:
        return pd.DataFrame()
    if df_csv.empty:
        # ainda assim retorna com status não encontrado
        df_base = df_xls.copy()
        c1 = _find_col_case_insensitive(df_base, "Guia TISS") or "Guia TISS"
        if c1 not in df_base.columns:
            df_base[c1] = None
        df_base["_key_guia"] = df_base[c1].apply(_norm_guia_tiss_tabela1)
        df_base["Status"] = "Não encontrado"
        return df_base

    col_xls = _find_col_case_insensitive(df_xls, "Guia TISS") or "Guia TISS"
    col_csv = _find_col_case_insensitive(df_csv, "Guia Tiss") or _find_col_case_insensitive(df_csv, "Guia TISS") or "Guia Tiss"

    dfx = df_xls.copy()
    if col_xls not in dfx.columns:
        dfx[col_xls] = None
    dfx["_key_guia"] = dfx[col_xls].apply(_norm_guia_tiss_tabela1)

    dfc = df_csv.copy()
    if col_csv not in dfc.columns:
        dfc[col_csv] = None
    dfc["_key_guia"] = dfc[col_csv].apply(_norm_guia_tiss_tabela2)

    # remove chaves vazias
    dfx = dfx[dfx["_key_guia"].astype(str).str.strip() != ""].copy()
    dfc = dfc[dfc["_key_guia"].astype(str).str.strip() != ""].copy()

    # se houver múltiplas linhas no CSV por chave, agregamos somando colunas numéricas
    num_cols = [c for c in dfc.columns if c != "_key_guia" and pd.api.types.is_numeric_dtype(dfc[c])]
    if num_cols:
        dfc_agg = dfc.groupby("_key_guia", as_index=False)[num_cols].sum()
        # pega a primeira ocorrência dos campos texto
        text_cols = [c for c in dfc.columns if c not in num_cols and c != "_key_guia"]
        if text_cols:
            dfc_first = dfc.groupby("_key_guia", as_index=False)[text_cols].first()
            dfc = pd.merge(dfc_first, dfc_agg, on="_key_guia", how="left")
        else:
            dfc = dfc_agg
    else:
        # dedupe simples
        dfc = dfc.drop_duplicates(subset=["_key_guia"]).copy()

    merged = dfx.merge(
        dfc,
        on="_key_guia",
        how="left",
        suffixes=("_XLS", "_CSV"),
        indicator=True
    )

    merged["Status"] = merged["_merge"].map({"both": "Encontrado", "left_only": "Não encontrado", "right_only": "(extra CSV)"})
    merged = merged.drop(columns=["_merge"])

    # Reorganiza colunas para ficar bom de analisar
    cols_front = []
    # status + chaves
    cols_front += ["Status", "_key_guia"]
    # guia original
    if col_xls in merged.columns:
        cols_front.append(col_xls)
    if col_csv in merged.columns:
        cols_front.append(col_csv)

    # tenta trazer campos úteis comuns da tabela1
    for c in ["Atendimento", "Nr.Conta", "Início", "Término", "Origem", "Valor"]:
        c_real = _find_col_case_insensitive(merged, c)
        if c_real and c_real not in cols_front:
            cols_front.append(c_real)

    # resto das colunas
    rest = [c for c in merged.columns if c not in cols_front]
    merged = merged[cols_front + rest]

    return merged.reset_index(drop=True)




def marcar_encontrados_csv(
    df_xls: pd.DataFrame,
    df_csv: pd.DataFrame,
    flag_col: str = "Encontrado",
    remessa_col_out: str = "Remessa (G5)",
    col_xls_nome: str = "Guia TISS",
    col_csv_nome: str = "Guia Tiss",
    col_csv_data_atend: str = "Data Atend",
    col_xls_termino: str = "Término",
) -> pd.DataFrame:
    """
    Marca no CSV (Tabela 2) os registros encontrados no XLS (Tabela 1) via Guia TISS normalizado.
    Também preenche uma coluna de remessa no CSV.

    Regra adicional (remessa exata):
    - Se a data do atendimento no CSV (col_csv_data_atend) bater com a data de término no XLS (col_xls_termino)
      para aquela mesma chave (Guia), então escolhe a remessa correspondente a essa data (um único número).
    - Se não bater, mantém o comportamento anterior: concatena remessas únicas (ex: "20735, 20736").
    """
    if df_csv is None:
        return pd.DataFrame()

    out = df_csv.copy()

    # garante coluna checkbox no começo
    if flag_col not in out.columns:
        out.insert(0, flag_col, False)
    else:
        out[flag_col] = out[flag_col].fillna(False).astype(bool)

    # garante coluna de remessa logo após a checkbox
    if remessa_col_out not in out.columns:
        out.insert(1, remessa_col_out, "")
    else:
        out[remessa_col_out] = out[remessa_col_out].fillna("").astype(str)

    if df_xls is None or df_xls.empty or out.empty:
        return out

    col_xls = _find_col_case_insensitive(df_xls, col_xls_nome) or col_xls_nome
    col_csv = (
        _find_col_case_insensitive(out, col_csv_nome)
        or _find_col_case_insensitive(out, "Guia TISS")
        or col_csv_nome
    )

    col_remessa_xls = _find_col_case_insensitive(df_xls, "Remessa (G5)") or "Remessa (G5)"
    col_termino_xls = _find_col_case_insensitive(df_xls, col_xls_termino) or col_xls_termino
    col_data_csv = _find_col_case_insensitive(out, col_csv_data_atend) or col_csv_data_atend

    if col_xls not in df_xls.columns or col_csv not in out.columns:
        return out

    # -------------------------
    # prepara tabela XLS com chave + (termino, remessa)
    # -------------------------
    dfx = df_xls.copy()

    if col_remessa_xls not in dfx.columns:
        dfx[col_remessa_xls] = ""
    if col_termino_xls not in dfx.columns:
        dfx[col_termino_xls] = None

    dfx["_key_guia"] = dfx[col_xls].apply(_norm_guia_tiss_tabela1).astype(str).str.strip()
    dfx = dfx[dfx["_key_guia"] != ""]

    # parse término
    dfx["_termino_dt"] = pd.to_datetime(dfx[col_termino_xls], errors="coerce", dayfirst=True).dt.date

    # normaliza remessa
    def _norm_rem(x):
        s = "" if x is None else str(x).strip()
        if s.lower() in ("nan", "none"):
            return ""
        return s

    dfx["_remessa_str"] = dfx[col_remessa_xls].apply(_norm_rem)
    dfx = dfx[dfx["_remessa_str"] != ""]

    # map chave -> lista de pares (termino_dt, remessa)
    pairs_map = (
        dfx.groupby("_key_guia")[["_termino_dt", "_remessa_str"]]
        .apply(lambda g: [(r["_termino_dt"], r["_remessa_str"]) for _, r in g.iterrows()])
        .to_dict()
    )

    # fallback: remessas únicas concatenadas
    rem_concat_map = {}
    for k, pairs in pairs_map.items():
        rems = sorted({p[1] for p in pairs if p[1]})
        rem_concat_map[k] = ", ".join(rems)

    keys_set = set(rem_concat_map.keys())

    # -------------------------
    # Regra NOVA: só marca/enche remessa se TISS + Data baterem
    # -------------------------
    keys_csv = out[col_csv].apply(_norm_guia_tiss_tabela2).astype(str).str.strip()

    # Data Atend (CSV) pode vir "07/10/2025 00:00:00" -> vira date(2025,10,7)
    if col_data_csv in out.columns:
        data_csv_dt = pd.to_datetime(out[col_data_csv], errors="coerce", dayfirst=True).dt.date
    else:
        data_csv_dt = pd.Series([None] * len(out))

    found_flags = []
    remessas_out = []

    for key, dt in zip(keys_csv.tolist(), data_csv_dt.tolist()):
        # precisa ter chave e data válida
        if not key or key not in pairs_map or dt is None:
            found_flags.append(False)
            remessas_out.append("")
            continue

        # procura remessa(s) no XLS cuja data Término == Data Atend
        matches = sorted({rem for (t, rem) in pairs_map[key] if t == dt and rem})

        if not matches:
            # TISS existe, mas data não bate -> não encontrado e sem remessa
            found_flags.append(False)
            remessas_out.append("")
            continue

        # achou pela data: encontrado e remessa exata (ou concat se tiver mais de uma)
        found_flags.append(True)
        remessas_out.append(matches[0] if len(matches) == 1 else ", ".join(matches))

    out[flag_col] = found_flags
    out[remessa_col_out] = remessas_out


    cols = [flag_col, remessa_col_out] + [c for c in out.columns if c not in (flag_col, remessa_col_out)]
    return out[cols]

